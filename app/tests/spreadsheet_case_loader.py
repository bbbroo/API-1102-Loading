from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import math
import re
from typing import Any, Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TESTING_SHEET = "Testing"
TESTING_WORKBOOKS = {
    "highway": REPO_ROOT / "Refs" / "API 1102 Highway_260606.xlsx",
    "railroad": REPO_ROOT / "Refs" / "API 1102 Railroad_260606.xlsx",
}
CACHED_FORMULA_MESSAGE = (
    "The Testing tab formula results are not cached. Open the workbook in Excel, "
    "recalculate, save, and rerun tests. openpyxl cannot evaluate Excel formulas by itself."
)
SPREADSHEET_ERROR_VALUES = {"#VALUE!", "#NAME?", "#REF!", "#N/A", "#DIV/0!", "#NUM!"}

SHARED_INPUT_ALIASES = {
    "nps": ["nps"],
    "wall_thickness": ["wall_thickness", "wall_thickness_in"],
    "pipe_specification": ["pipe_specification"],
    "pipe_grade": ["pipe_grade"],
    "pipeline_location": ["pipeline_location"],
    "class_location": ["class_location"],
    "weld_seam_type": ["weld_seam_type"],
    "pipe_material": ["pipe_material"],
    "operating_temperature": ["operating_temperature", "operating_temp", "operating_temp_f"],
    "installation_temperature": ["installation_temperature", "installation_temp", "installation_temp_f"],
    "operating_pressure": ["operating_pressure", "operating_pressure_psig"],
    "soil_type": ["soil_type"],
    "soil_unit_weight": ["soil_unit_weight", "soil_unit_weight_pcf"],
    "cover_depth": ["cover_depth", "cover_depth_ft"],
    "bored_diameter": ["bored_diameter", "bored_diameter_in"],
}

HIGHWAY_INPUT_ALIASES = {
    "pavement_type": ["pavement_type"],
    "axle_configuration": ["axle_configuration"],
}

RAILROAD_INPUT_ALIASES = {
    "number_of_tracks": ["number_of_tracks"],
    "surface_pressure": ["surface_pressure", "surface_pressure_psi"],
}


@dataclass(frozen=True)
class OutputMapping:
    output_name: str
    aliases: tuple[str, ...]
    source: str
    app_key: str
    abs_tolerance: float
    rel_tolerance: float


@dataclass(frozen=True)
class SpreadsheetCase:
    workbook_path: Path
    sheet_name: str
    excel_row: int
    case_id: str
    mode: str
    row: dict[str, Any]
    raw_headers: dict[str, str]
    shared_inputs: dict[str, Any]
    mode_inputs: dict[str, Any]


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-\/()]+", "_", text)
    text = re.sub(r"[^a-z0-9_]+", "", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def locate_testing_workbook(mode: str) -> Path:
    preferred = TESTING_WORKBOOKS.get(mode.lower())
    if preferred and preferred.exists():
        return preferred
    matches: list[Path] = []
    for path in REPO_ROOT.rglob("*.xlsx"):
        if ".bak_" in path.name:
            continue
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            if TESTING_SHEET not in workbook.sheetnames:
                continue
            sheet = workbook[TESTING_SHEET]
            mode_value = sheet["C2"].value
            if str(mode_value).strip().lower() == mode.lower() or mode.lower() in path.name.lower():
                matches.append(path)
        finally:
            workbook.close()
    if not matches:
        raise AssertionError(f"No {mode} workbook with a Testing worksheet was found.")
    if len(matches) > 1:
        names = ", ".join(str(path) for path in matches)
        raise AssertionError(f"Multiple {mode} Testing workbooks were found: {names}")
    return matches[0]


def load_highway_testing_cases() -> list[SpreadsheetCase]:
    return load_testing_cases("Highway", 168, HIGHWAY_INPUT_ALIASES)


def load_railroad_testing_cases() -> list[SpreadsheetCase]:
    return load_testing_cases("Railroad", 144, RAILROAD_INPUT_ALIASES)


def load_testing_cases(mode: str, expected_count: int, mode_aliases: dict[str, list[str]]) -> list[SpreadsheetCase]:
    workbook_path = locate_testing_workbook(mode)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if TESTING_SHEET not in workbook.sheetnames:
            raise AssertionError(f"Workbook {workbook_path.name} does not contain a Testing worksheet.")
        sheet = workbook[TESTING_SHEET]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(value) for value in header_cells]
        raw_headers = {header: str(raw).strip() for header, raw in zip(headers, header_cells) if header}
        cases: list[SpreadsheetCase] = []
        for excel_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value in (None, "") for value in values):
                continue
            row = {header: value for header, value in zip(headers, values) if header}
            shared_inputs = extract_inputs(row, SHARED_INPUT_ALIASES)
            mode_inputs = extract_inputs(row, mode_aliases)
            case_id = str(first_present(row, ["case_id", "case_index"]) or excel_row)
            cases.append(
                SpreadsheetCase(
                    workbook_path=workbook_path,
                    sheet_name=TESTING_SHEET,
                    excel_row=excel_row,
                    case_id=case_id,
                    mode=mode,
                    row=row,
                    raw_headers=raw_headers,
                    shared_inputs=shared_inputs,
                    mode_inputs=mode_inputs,
                )
            )
        if len(cases) != expected_count:
            raise AssertionError(f"{mode} Testing tab row count is {len(cases)}, expected {expected_count}. Workbook: {workbook_path}")
        return cases
    finally:
        workbook.close()


def extract_inputs(row: dict[str, Any], aliases: dict[str, list[str]]) -> dict[str, Any]:
    inputs: dict[str, Any] = {}
    for app_key, names in aliases.items():
        value = first_present(row, names)
        if value not in (None, ""):
            inputs[app_key] = value
    return inputs


def first_present(row: dict[str, Any], aliases: Iterable[str]) -> Any:
    for alias in aliases:
        key = normalize_header(alias)
        if key in row:
            return row[key]
    return None


def output_value(case: SpreadsheetCase, mapping: OutputMapping) -> Any:
    return first_present(case.row, mapping.aliases)


def output_header_present(case: SpreadsheetCase, mapping: OutputMapping) -> bool:
    return any(normalize_header(alias) in case.row for alias in mapping.aliases)


def assert_required_headers(cases: list[SpreadsheetCase], aliases: dict[str, list[str]], output_mappings: list[OutputMapping]) -> None:
    case = cases[0]
    missing_inputs = [name for name, options in aliases.items() if not any(normalize_header(option) in case.row for option in options)]
    missing_outputs = [mapping.output_name for mapping in output_mappings if not output_header_present(case, mapping)]
    if missing_inputs or missing_outputs:
        raise AssertionError(
            f"Testing tab schema mismatch in {case.workbook_path.name}. "
            f"Missing input headers: {missing_inputs}. Missing output headers: {missing_outputs}."
        )


def assert_no_spreadsheet_errors(cases: list[SpreadsheetCase], output_mappings: list[OutputMapping]) -> None:
    for case in cases:
        for mapping in output_mappings:
            value = output_value(case, mapping)
            if isinstance(value, str) and value.strip() in SPREADSHEET_ERROR_VALUES:
                raise AssertionError(
                    f"Spreadsheet error value found in Testing tab output cell. Workbook: {case.workbook_path.name}; "
                    f"Sheet: {case.sheet_name}; Excel row: {case.excel_row}; Case ID: {case.case_id}; "
                    f"Output: {mapping.output_name}; Value: {value}"
                )


def assert_cached_outputs_available(cases: list[SpreadsheetCase], output_mappings: list[OutputMapping]) -> None:
    for case in cases:
        for mapping in output_mappings:
            value = output_value(case, mapping)
            if value not in (None, ""):
                if isinstance(value, str) and value.strip().startswith("="):
                    raise AssertionError(CACHED_FORMULA_MESSAGE)
                return
    raise AssertionError(CACHED_FORMULA_MESSAGE)


def check_map(result) -> dict[str, Any]:
    return {check.name: check for check in result.checks}


def app_output_value(result, mapping: OutputMapping) -> float:
    if mapping.source == "intermediate":
        return float(result.intermediate_values[mapping.app_key])
    checks = check_map(result)
    if mapping.source == "check_calculated":
        return float(checks[mapping.app_key].calculated_psi)
    if mapping.source == "check_allowable":
        return float(checks[mapping.app_key].allowable_psi)
    raise AssertionError(f"Unknown output mapping source: {mapping.source}")


def assert_outputs_match(case: SpreadsheetCase, result, output_mappings: list[OutputMapping]) -> list[str]:
    failures = []
    for mapping in output_mappings:
        expected = output_value(case, mapping)
        if expected in (None, ""):
            continue
        if isinstance(expected, str) and expected.strip().startswith("="):
            raise AssertionError(CACHED_FORMULA_MESSAGE)
        try:
            expected_value = float(expected)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(expected_value):
            continue
        actual = app_output_value(result, mapping)
        delta = actual - expected_value
        abs_ok = abs(delta) <= mapping.abs_tolerance
        rel_ok = abs(delta) / max(abs(expected_value), 1.0) <= mapping.rel_tolerance
        if not (abs_ok or rel_ok):
            failures.append(format_mismatch(case, mapping, expected_value, actual, delta))
    return failures


def format_mismatch(case: SpreadsheetCase, mapping: OutputMapping, expected: float, actual: float, delta: float) -> str:
    input_lines = {
        **case.shared_inputs,
        **case.mode_inputs,
    }
    interesting = [
        "nps",
        "wall_thickness",
        "cover_depth",
        "bored_diameter",
        "operating_pressure",
        "soil_type",
        "pavement_type",
        "axle_configuration",
        "surface_pressure",
        "number_of_tracks",
    ]
    inputs_text = "\n".join(f"  {key}={input_lines.get(key)}" for key in interesting if key in input_lines)
    return (
        "Spreadsheet parity mismatch:\n"
        f"Workbook: {case.workbook_path.name}\n"
        f"Sheet: {case.sheet_name}\n"
        f"Excel row: {case.excel_row}\n"
        f"Case ID: {case.case_id}\n"
        f"Mode: {case.mode}\n"
        f"Output: {mapping.output_name}\n"
        f"Spreadsheet value: {expected}\n"
        f"App value: {actual}\n"
        f"Delta: {delta}\n"
        f"Abs tolerance: {mapping.abs_tolerance}\n"
        f"Rel tolerance: {mapping.rel_tolerance}\n"
        "Inputs:\n"
        f"{inputs_text}"
    )


def intermediate_output(name: str, app_key: str, *aliases: str, factor: bool = False) -> OutputMapping:
    return OutputMapping(
        output_name=name,
        aliases=tuple(aliases or (name,)),
        source="intermediate",
        app_key=app_key,
        abs_tolerance=0.001 if factor else 1.0,
        rel_tolerance=0.001,
    )


def check_calculated_output(name: str, check_name: str, *aliases: str) -> OutputMapping:
    return OutputMapping(name, tuple(aliases or (name,)), "check_calculated", check_name, 1.0, 0.001)


def check_allowable_output(name: str, check_name: str, *aliases: str) -> OutputMapping:
    return OutputMapping(name, tuple(aliases or (name,)), "check_allowable", check_name, 1.0, 0.001)


COMMON_OUTPUTS = [
    intermediate_output("SHi", "SHi", "SHi_psi", "SHi Barlow psi", "SHi_Barlow_psi"),
    intermediate_output("SHi_internal", "SHi_internal", "SHi_Internal_psi"),
    intermediate_output("SHe", "SHe", "SHe_psi"),
    intermediate_output("S1", "S1", "S1_psi"),
    intermediate_output("S2", "S2", "S2_psi"),
    intermediate_output("S3", "S3", "S3_psi"),
    intermediate_output("Seff", "Seff", "Seff_psi"),
    intermediate_output("allowable_hoop", "allowable_hoop", "Allowable_Hoop_psi"),
    intermediate_output("allowable_effective", "allowable_effective", "Allowable_Effective_psi"),
    intermediate_output("fatigue_girth", "fatigue_girth", "Fatigue_Girth_psi"),
    intermediate_output("fatigue_longitudinal", "fatigue_longitudinal", "Fatigue_Longitudinal_psi"),
    intermediate_output("allowable_girth", "allowable_girth", "Allowable_Girth_psi"),
    intermediate_output("allowable_longitudinal", "allowable_longitudinal", "Allowable_Longitudinal_psi"),
    intermediate_output("Khe", "Khe", factor=True),
    intermediate_output("Be", "Be", factor=True),
    intermediate_output("Ee", "Ee", factor=True),
    intermediate_output("Fi", "Fi", factor=True),
    check_calculated_output("Barlow_Calc", "Barlow Stress", "Barlow_Calc", "Barlow_Calc_psi"),
    check_allowable_output("Barlow_Allow", "Barlow Stress", "Barlow_Allow", "Barlow_Allow_psi"),
    check_calculated_output("Effective_Calc", "Effective Stress", "Effective_Calc", "Effective_Calc_psi"),
    check_allowable_output("Effective_Allow", "Effective Stress", "Effective_Allow", "Effective_Allow_psi"),
    check_calculated_output("Girth_Calc", "Girth Weld Stress", "Girth_Calc", "Girth_Calc_psi"),
    check_allowable_output("Girth_Allow", "Girth Weld Stress", "Girth_Allow", "Girth_Allow_psi"),
    check_calculated_output("Longitudinal_Calc", "Longitudinal Weld Stress", "Longitudinal_Calc", "Longitudinal_Calc_psi"),
    check_allowable_output("Longitudinal_Allow", "Longitudinal Weld Stress", "Longitudinal_Allow", "Longitudinal_Allow_psi"),
]

HIGHWAY_OUTPUT_MAPPINGS = [
    *COMMON_OUTPUTS,
    intermediate_output("SHh", "SHh", "SHh_psi"),
    intermediate_output("SLh", "SLh", "SLh_psi"),
    intermediate_output("KHh", "KHh", factor=True),
    intermediate_output("GHh", "GHh", factor=True),
    intermediate_output("KLh", "KLh", factor=True),
    intermediate_output("GLh", "GLh", factor=True),
    intermediate_output("R", "R", factor=True),
    intermediate_output("L", "L", factor=True),
]

RAILROAD_OUTPUT_MAPPINGS = [
    *COMMON_OUTPUTS,
    intermediate_output("SHr", "SHr", "SHr_psi"),
    intermediate_output("SLr", "SLr", "SLr_psi"),
    intermediate_output("Nh", "Nh", factor=True),
    intermediate_output("KHr", "KHr", factor=True),
    intermediate_output("GHr", "GHr", factor=True),
    intermediate_output("NL", "NL", factor=True),
    intermediate_output("KLr", "KLr", factor=True),
    intermediate_output("GLr", "GLr", factor=True),
]
