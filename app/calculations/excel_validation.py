from __future__ import annotations

import importlib.util
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from app.calculations.highway import calculate_highway
from app.calculations.railroad import calculate_railroad

REPO_ROOT = Path(__file__).resolve().parents[2]
REFS = REPO_ROOT / "Refs"
HIGHWAY_WORKBOOK = REFS / "Copy of API 1102 Highway.xlsx"
RAILROAD_WORKBOOK = REFS / "Copy of API 1102 Railroad.xlsx"
STRESS_ABS_TOLERANCE_PSI = 0.5
RELATIVE_TOLERANCE = 0.001
HIGHWAY_MAPPING = {"SHi": "C44", "SHe": "C50", "SHh": "C58", "SLh": "C61", "Seff": "C67", "Barlow": "C78", "Effective": "C79", "Girth": "C80", "Longitudinal": "C81"}
RAILROAD_MAPPING = {"SHi": "C43", "SHe": "C49", "SHr": "C56", "SLr": "C60", "Seff": "C66", "Barlow": "C78", "Effective": "C79", "Girth": "C80", "Longitudinal": "C81"}


def read_cells(path: Path, sheet: str, mapping: dict[str, str]) -> dict[str, float | str]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet]
    return {name: ws[cell].value for name, cell in mapping.items()}


def compare_default_highway() -> dict[str, dict[str, float]]:
    expected = read_cells(HIGHWAY_WORKBOOK, "Highway Loading", HIGHWAY_MAPPING)
    return diff(expected, highway_actuals(), "Highway Loading", HIGHWAY_MAPPING)


def compare_default_railroad() -> dict[str, dict[str, float]]:
    expected = read_cells(RAILROAD_WORKBOOK, "Railroad Loading", RAILROAD_MAPPING)
    return diff(expected, railroad_actuals(), "Railroad Loading", RAILROAD_MAPPING)


def compare_highway_case(shared_inputs: dict | None = None, highway_inputs: dict | None = None) -> dict[str, dict[str, float]]:
    expected = recalculate_case(HIGHWAY_WORKBOOK, "Highway Loading", highway_cell_updates(shared_inputs or {}, highway_inputs or {}), HIGHWAY_MAPPING)
    return diff(expected, highway_actuals(shared_inputs, highway_inputs), "Highway Loading", HIGHWAY_MAPPING)


def compare_railroad_case(shared_inputs: dict | None = None, railroad_inputs: dict | None = None) -> dict[str, dict[str, float]]:
    expected = recalculate_case(RAILROAD_WORKBOOK, "Railroad Loading", railroad_cell_updates(shared_inputs or {}, railroad_inputs or {}), RAILROAD_MAPPING)
    return diff(expected, railroad_actuals(shared_inputs, railroad_inputs), "Railroad Loading", RAILROAD_MAPPING)


def highway_actuals(shared_inputs: dict | None = None, highway_inputs: dict | None = None) -> dict[str, float]:
    result = calculate_highway(shared_inputs, highway_inputs)
    actual = result.intermediate_values
    checks = {c.name: c.calculated_psi for c in result.checks}
    return {"SHi": actual["SHi"], "SHe": actual["SHe"], "SHh": actual["SHh"], "SLh": actual["SLh"], "Seff": actual["Seff"], "Barlow": checks["Barlow Stress"], "Effective": checks["Effective Stress"], "Girth": checks["Girth Weld Stress"], "Longitudinal": checks["Longitudinal Weld Stress"]}


def railroad_actuals(shared_inputs: dict | None = None, railroad_inputs: dict | None = None) -> dict[str, float]:
    result = calculate_railroad(shared_inputs, railroad_inputs)
    actual = result.intermediate_values
    checks = {c.name: c.calculated_psi for c in result.checks}
    return {"SHi": actual["SHi"], "SHe": actual["SHe"], "SHr": actual["SHr"], "SLr": actual["SLr"], "Seff": actual["Seff"], "Barlow": checks["Barlow Stress"], "Effective": checks["Effective Stress"], "Girth": checks["Girth Weld Stress"], "Longitudinal": checks["Longitudinal Weld Stress"]}


def has_excel_recalculator() -> bool:
    try:
        return importlib.util.find_spec("win32com.client") is not None
    except ModuleNotFoundError:
        return False


def recalculate_case(path: Path, sheet: str, updates: dict[str, object], mapping: dict[str, str]) -> dict[str, float | str]:
    if not has_excel_recalculator():
        raise RuntimeError("pywin32 is required for edited workbook recalculation tests.")
    with tempfile.TemporaryDirectory() as tmpdir:
        target = Path(tmpdir) / path.name
        shutil.copy2(path, target)
        wb = load_workbook(target, data_only=False, read_only=False)
        ws = wb[sheet]
        for cell, value in updates.items():
            ws[cell] = value
        wb.save(target)
        recalculate_workbook(target)
        return read_cells(target, sheet, mapping)


def recalculate_workbook(path: Path) -> None:
    import win32com.client  # type: ignore

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(path.resolve()))
        workbook.ForceFullCalculation = True
        workbook.RefreshAll()
        excel.CalculateFullRebuild()
        workbook.Save()
        workbook.Close(SaveChanges=True)
    finally:
        excel.Quit()


def highway_cell_updates(shared: dict, highway: dict) -> dict[str, object]:
    return shared_cell_updates(shared, installation_cell="C39") | {
        "C36": highway.get("pavement_type", "Flexible"),
        "C37": highway.get("axle_configuration", "Tandem Axle"),
    }


def railroad_cell_updates(shared: dict, railroad: dict) -> dict[str, object]:
    return shared_cell_updates(shared, installation_cell="C36") | {
        "C37": int(railroad.get("number_of_tracks", 2)),
        "C38": railroad.get("surface_pressure", 13.9),
    }


def shared_cell_updates(shared: dict, installation_cell: str) -> dict[str, object]:
    updates: dict[str, object] = {}
    cell_map = {
        "nps": "C12",
        "wall_thickness": "C14",
        "pipe_specification": "C15",
        "pipe_grade": "C16",
        "pipeline_location": "C18",
        "class_location": "C19",
        "weld_seam_type": "C21",
        "pipe_material": "C23",
        "operating_temperature": "C27",
        "operating_pressure": "C29",
        "soil_type": "C30",
        "soil_unit_weight": "C33",
        "cover_depth": "C34",
        "bored_diameter": "C35",
        "installation_temperature": installation_cell,
    }
    for key, cell in cell_map.items():
        if key in shared:
            value = shared[key]
            updates[cell] = nps_for_excel(value) if key == "nps" else class_for_excel(value) if key == "class_location" else value
    return updates


def nps_for_excel(value: object) -> object:
    text = str(value)
    return int(text) if text.isdigit() else text


def class_for_excel(value: object) -> object:
    text = str(value)
    return int(text) if text.isdigit() else value


def diff(expected: dict, actual: dict, sheet: str = "", mapping: dict[str, str] | None = None) -> dict[str, dict[str, float | str | bool]]:
    out = {}
    for key, exp in expected.items():
        act = actual[key]
        expected_value = float(exp)
        actual_value = float(act)
        delta = actual_value - expected_value
        relative_delta = abs(delta) / abs(expected_value) if expected_value else abs(delta)
        out[key] = {
            "sheet": sheet,
            "cell": (mapping or {}).get(key, ""),
            "expected": expected_value,
            "actual": actual_value,
            "delta": delta,
            "relative_delta": relative_delta,
            "within_tolerance": abs(delta) <= STRESS_ABS_TOLERANCE_PSI or relative_delta <= RELATIVE_TOLERANCE,
        }
    return out


def assert_within_tolerance(report: dict[str, dict[str, float | str | bool]]) -> None:
    failures = {name: data for name, data in report.items() if not data["within_tolerance"]}
    if failures:
        lines = [
            f"{name} {data['sheet']}!{data['cell']}: expected={data['expected']} actual={data['actual']} delta={data['delta']}"
            for name, data in failures.items()
        ]
        raise AssertionError("Excel validation mismatches exceeded tolerance:\n" + "\n".join(lines))
