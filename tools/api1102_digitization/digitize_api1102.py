from __future__ import annotations

import csv
import json
import math
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import fitz
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.api1102_digitization.interpolation import DigitizedRangeError, linear_interpolate

PDF_PATH = Path(r"C:\Users\brook\Downloads\API 1102.pdf")
OUTPUT_DIR = ROOT / "Refs" / "digitized_api_1102"
OVERLAY_DIR = OUTPUT_DIR / "overlays"
PAGE_IMAGE_DIR = OUTPUT_DIR / "source_page_images"
CLEAN_GRAPH_DIR = OUTPUT_DIR / "graph_underlays"
RAILROAD_WORKBOOK = ROOT / "Refs" / "API 1102 Railroad_260606.xlsx"
HIGHWAY_WORKBOOK = ROOT / "Refs" / "API 1102 Highway_260606.xlsx"

DIGITIZATION_METHOD = "PDF vector path extraction; piecewise labeled-tick axis calibration; no OCR"
CURVE_SAMPLE_COUNT = 31
BEZIER_STEPS = 160
LINE_STEPS = 12
PAGE_COORD_TOLERANCE = 1.5
FIRST_PASS_ENDPOINT_SPAN_SUMMARY = {
    "calibration": "endpoint-span affine calibration superseded by piecewise labeled-tick calibration",
    "comparison_status": {"PASS": 413, "REVIEW": 156, "FAIL": 143},
    "failed_factors": ["GLh", "GLr", "KHe", "KHh", "KLr"],
}


@dataclass(frozen=True)
class DrawingPart:
    drawing_id: int
    item_indices: tuple[int, ...] | None = None


@dataclass(frozen=True)
class CurveSpec:
    curve_name: str
    parts: tuple[DrawingPart, ...]
    tail_drawing_id: int | None = None
    notes: str = ""


@dataclass(frozen=True)
class FigureSpec:
    figure: str
    factor: str
    filename: str
    sheet_name: str
    pdf_page: int
    api_page: int
    frame: tuple[float, float, float, float]
    axis_x: tuple[float, float]
    axis_y: tuple[float, float]
    x_units: str
    y_units: str
    curves: tuple[CurveSpec, ...]
    orientation: str = "normal"
    uniform_count: int = CURVE_SAMPLE_COUNT
    graph_note: str = ""


@dataclass(frozen=True)
class TablePair:
    x_col: int
    y_col: int
    curve_name: str


@dataclass(frozen=True)
class SpreadsheetSpec:
    workbook_label: str
    path: Path
    figure: str
    factor: str
    row_start: int
    row_end: int
    pairs: tuple[TablePair, ...]


@dataclass
class DensePoint:
    x: float
    y: float
    page_x: float
    page_y: float


@dataclass
class DigitizedRecord:
    figure: str
    factor: str
    curve_name: str
    x_value: float
    x_units: str
    y_value: float
    y_units: str
    source_page: str
    digitization_method: str
    notes: str
    point_type: str
    page_x: float
    page_y: float


class CalibrationRangeError(ValueError):
    """Raised when a page or data coordinate is outside a calibrated tick range."""


@dataclass(frozen=True)
class AxisTick:
    value: float
    page_coord: float
    label: str


@dataclass(frozen=True)
class PiecewiseAxisCalibration:
    axis_name: str
    page_coordinate: str
    source_label: str
    ignored_secondary_axis: str
    ticks: tuple[AxisTick, ...]
    coordinate_tolerance: float = PAGE_COORD_TOLERANCE
    value_tolerance: float = 1e-10

    def __post_init__(self) -> None:
        if len(self.ticks) < 2:
            raise ValueError(f"{self.axis_name} requires at least two calibration ticks.")
        page_coords = [tick.page_coord for tick in self.ticks]
        values = [tick.value for tick in self.ticks]
        if len(set(page_coords)) != len(page_coords):
            raise ValueError(f"{self.axis_name} has duplicate page coordinates.")
        if len(set(values)) != len(values):
            raise ValueError(f"{self.axis_name} has duplicate tick values.")

    @property
    def value_range(self) -> tuple[float, float]:
        values = [tick.value for tick in self.ticks]
        return min(values), max(values)

    @property
    def page_range(self) -> tuple[float, float]:
        coords = [tick.page_coord for tick in self.ticks]
        return min(coords), max(coords)

    def page_to_value(self, page_coord: float) -> float:
        return self._interpolate(
            page_coord,
            sorted((tick.page_coord, tick.value) for tick in self.ticks),
            tolerance=self.coordinate_tolerance,
            input_name="page coordinate",
        )

    def value_to_page(self, value: float) -> float:
        return self._interpolate(
            value,
            sorted((tick.value, tick.page_coord) for tick in self.ticks),
            tolerance=self.value_tolerance,
            input_name="axis value",
        )

    def _interpolate(
        self,
        input_value: float,
        ordered_pairs: list[tuple[float, float]],
        *,
        tolerance: float,
        input_name: str,
    ) -> float:
        x = float(input_value)
        xmin = ordered_pairs[0][0]
        xmax = ordered_pairs[-1][0]
        if x < xmin - tolerance or x > xmax + tolerance:
            raise CalibrationRangeError(
                f"{self.axis_name} {input_name} {x} is outside calibrated tick range {xmin} to {xmax}."
            )
        if x < xmin:
            x = xmin
        if x > xmax:
            x = xmax
        for (x0, y0), (x1, y1) in zip(ordered_pairs, ordered_pairs[1:]):
            if x0 - tolerance <= x <= x1 + tolerance:
                if abs(x1 - x0) <= 1e-12:
                    return y0
                fraction = (x - x0) / (x1 - x0)
                return y0 + fraction * (y1 - y0)
        raise CalibrationRangeError(f"{self.axis_name} {input_name} {x} could not be bracketed by calibration ticks.")

    def residuals(self) -> dict[str, float]:
        page_residuals = [abs(self.value_to_page(tick.value) - tick.page_coord) for tick in self.ticks]
        value_residuals = [abs(self.page_to_value(tick.page_coord) - tick.value) for tick in self.ticks]
        return {
            "max_page_coord_residual": max(page_residuals, default=0.0),
            "max_value_residual": max(value_residuals, default=0.0),
        }

    def to_metadata(self) -> dict[str, Any]:
        residuals = self.residuals()
        return {
            "axis_name": self.axis_name,
            "calibration_method": "piecewise_linear_labeled_ticks",
            "page_coordinate": self.page_coordinate,
            "source_label": self.source_label,
            "ignored_secondary_axis": self.ignored_secondary_axis,
            "ticks": [
                {"value": tick.value, "page_coord": tick.page_coord, "label": tick.label}
                for tick in self.ticks
            ],
            **residuals,
        }


FIGURES: dict[str, FigureSpec] = {
    "03": FigureSpec(
        figure="Figure 3",
        factor="KHe",
        filename="figure_03_KHe.csv",
        sheet_name="Fig03_KHe",
        pdf_page=19,
        api_page=13,
        frame=(112.45, 102.35, 551.59, 330.57),
        axis_x=(0.0, 0.08),
        axis_y=(0.0, 12000.0),
        x_units="tw/D",
        y_units="dimensionless",
        curves=(
            CurveSpec("E_prime_ksi=0.2", (DrawingPart(60),), tail_drawing_id=57, notes="Shared common tail after curve convergence."),
            CurveSpec("E_prime_ksi=0.5", (DrawingPart(59),), tail_drawing_id=57, notes="Shared common tail after curve convergence."),
            CurveSpec("E_prime_ksi=1.0", (DrawingPart(58),), tail_drawing_id=57, notes="Shared common tail after curve convergence."),
            CurveSpec("E_prime_ksi=2.0", (DrawingPart(57),)),
        ),
        graph_note="KHe curves converge; the single visible common tail is reused for upper E' curves after their unique strokes end.",
    ),
    "04": FigureSpec(
        figure="Figure 4",
        factor="Be",
        filename="figure_04_Be.csv",
        sheet_name="Fig04_Be",
        pdf_page=19,
        api_page=13,
        frame=(161.94, 422.13, 487.56, 665.24),
        axis_x=(0.0, 32.0),
        axis_y=(0.0, 1.5),
        x_units="H/Bd",
        y_units="dimensionless",
        curves=(
            CurveSpec("soil_type=A", (DrawingPart(404),)),
            CurveSpec("soil_type=B", (DrawingPart(403),)),
        ),
    ),
    "05": FigureSpec(
        figure="Figure 5",
        factor="Ee",
        filename="figure_05_Ee.csv",
        sheet_name="Fig05_Ee",
        pdf_page=20,
        api_page=14,
        frame=(155.79, 136.97, 479.21, 356.43),
        axis_x=(1.0, 1.3),
        axis_y=(0.8, 1.4),
        x_units="Bd/D",
        y_units="dimensionless",
        curves=(CurveSpec("Ee", (DrawingPart(5),)),),
    ),
    "07": FigureSpec(
        figure="Figure 7",
        factor="Fi",
        filename="figure_07_Fi.csv",
        sheet_name="Fig07_Fi",
        pdf_page=22,
        api_page=16,
        frame=(195.69, 128.27, 406.35, 447.28),
        axis_x=(1.0, 2.0),
        axis_y=(0.0, 30.0),
        x_units="ft",
        y_units="dimensionless",
        curves=(
            CurveSpec("crossing=highway", (DrawingPart(14),)),
            CurveSpec("crossing=railroad", (DrawingPart(15),)),
        ),
        orientation="depth_on_y",
        uniform_count=19,
        graph_note="Graph plots Fi on the horizontal axis and depth on the vertical axis; CSV normalizes x as depth H.",
    ),
    "08": FigureSpec(
        figure="Figure 8",
        factor="KHr",
        filename="figure_08_KHr.csv",
        sheet_name="Fig08_KHr",
        pdf_page=23,
        api_page=17,
        frame=(161.41, 107.62, 475.76, 303.96),
        axis_x=(0.0, 0.08),
        axis_y=(0.0, 500.0),
        x_units="tw/D",
        y_units="dimensionless",
        curves=(
            CurveSpec("Er_ksi=5", (DrawingPart(59),)),
            CurveSpec("Er_ksi=10", (DrawingPart(60),)),
            CurveSpec("Er_ksi=20", (DrawingPart(61),)),
        ),
    ),
    "09": FigureSpec(
        figure="Figure 9",
        factor="GHr",
        filename="figure_09_GHr.csv",
        sheet_name="Fig09_GHr",
        pdf_page=24,
        api_page=18,
        frame=(171.59, 151.51, 471.29, 366.03),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 1.25),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(53),)),
            CurveSpec("H_ft=10", (DrawingPart(51),)),
            CurveSpec("H_ft=14", (DrawingPart(52),)),
        ),
    ),
    "10": FigureSpec(
        figure="Figure 10",
        factor="NH",
        filename="figure_10_NH.csv",
        sheet_name="Fig10_NH",
        pdf_page=25,
        api_page=19,
        frame=(163.51, 127.85, 477.62, 360.38),
        axis_x=(0.0, 42.0),
        axis_y=(0.5, 2.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(65),)),
            CurveSpec("H_ft=10", (DrawingPart(64),)),
            CurveSpec("H_ft=14", (DrawingPart(63),)),
        ),
    ),
    "11": FigureSpec(
        figure="Figure 11",
        factor="KLr",
        filename="figure_11_KLr.csv",
        sheet_name="Fig11_KLr",
        pdf_page=25,
        api_page=19,
        frame=(167.15, 433.41, 497.72, 680.37),
        axis_x=(0.0, 0.08),
        axis_y=(0.0, 600.0),
        x_units="tw/D",
        y_units="dimensionless",
        curves=(
            CurveSpec("Er_ksi=5", (DrawingPart(134),)),
            CurveSpec("Er_ksi=10", (DrawingPart(135),)),
            CurveSpec("Er_ksi=20", (DrawingPart(136),)),
        ),
    ),
    "12": FigureSpec(
        figure="Figure 12",
        factor="GLr",
        filename="figure_12_GLr.csv",
        sheet_name="Fig12_GLr",
        pdf_page=26,
        api_page=20,
        frame=(180.20, 132.66, 480.18, 347.99),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 2.5),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(60),)),
            CurveSpec("H_ft=10", (DrawingPart(61),)),
            CurveSpec("H_ft=14", (DrawingPart(62),)),
        ),
    ),
    "13": FigureSpec(
        figure="Figure 13",
        factor="NL",
        filename="figure_13_NL.csv",
        sheet_name="Fig13_NL",
        pdf_page=26,
        api_page=20,
        frame=(158.98, 446.18, 471.62, 687.91),
        axis_x=(0.0, 42.0),
        axis_y=(0.5, 2.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(125),)),
            CurveSpec("H_ft=10", (DrawingPart(124),)),
            CurveSpec("H_ft=14", (DrawingPart(123),)),
        ),
    ),
    "14": FigureSpec(
        figure="Figure 14",
        factor="KHh",
        filename="figure_14_KHh.csv",
        sheet_name="Fig14_KHh",
        pdf_page=27,
        api_page=21,
        frame=(175.96, 182.42, 488.08, 377.81),
        axis_x=(0.0, 0.08),
        axis_y=(0.0, 25.0),
        x_units="tw/D",
        y_units="dimensionless",
        curves=(
            CurveSpec("Er_ksi=5", (DrawingPart(60),)),
            CurveSpec("Er_ksi=10", (DrawingPart(61),)),
            CurveSpec("Er_ksi=20", (DrawingPart(62),)),
        ),
    ),
    "15": FigureSpec(
        figure="Figure 15",
        factor="GHh",
        filename="figure_15_GHh.csv",
        sheet_name="Fig15_GHh",
        pdf_page=28,
        api_page=22,
        frame=(158.93, 146.10, 472.18, 369.40),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 2.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=3_to_4", (DrawingPart(52),)),
            CurveSpec("H_ft=6", (DrawingPart(53),)),
            CurveSpec("H_ft=8", (DrawingPart(51),)),
            CurveSpec("H_ft=10", (DrawingPart(41),)),
        ),
    ),
    "16": FigureSpec(
        figure="Figure 16",
        factor="KLh",
        filename="figure_16_KLh.csv",
        sheet_name="Fig16_KLh",
        pdf_page=29,
        api_page=23,
        frame=(170.23, 112.66, 485.64, 310.10),
        axis_x=(0.0, 0.08),
        axis_y=(0.0, 25.0),
        x_units="tw/D",
        y_units="dimensionless",
        curves=(
            CurveSpec("Er_ksi=5", (DrawingPart(14),)),
            CurveSpec("Er_ksi=10", (DrawingPart(12),)),
            CurveSpec("Er_ksi=20", (DrawingPart(13),)),
        ),
    ),
    "17": FigureSpec(
        figure="Figure 17",
        factor="GLh",
        filename="figure_17_GLh.csv",
        sheet_name="Fig17_GLh",
        pdf_page=29,
        api_page=23,
        frame=(166.19, 434.62, 479.18, 657.73),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 3.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=3_to_4", (DrawingPart(65),)),
            CurveSpec("H_ft=6", (DrawingPart(66),)),
            CurveSpec("H_ft=8", (DrawingPart(67),)),
            CurveSpec("H_ft=10", (DrawingPart(68),)),
        ),
    ),
    "18A": FigureSpec(
        figure="Figure 18-A",
        factor="RF",
        filename="figure_18A_RF.csv",
        sheet_name="Fig18A_RF",
        pdf_page=34,
        api_page=28,
        frame=(154.29, 128.59, 478.26, 357.95),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 1.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(54, (2, 3)),)),
            CurveSpec("H_ft=10", (DrawingPart(54, (0, 1)),)),
            CurveSpec("H_ft=14", (DrawingPart(53),)),
        ),
        uniform_count=17,
        graph_note="For 5 ft <= LG < 10 ft.",
    ),
    "18B": FigureSpec(
        figure="Figure 18-B",
        factor="RF",
        filename="figure_18B_RF.csv",
        sheet_name="Fig18B_RF",
        pdf_page=34,
        api_page=28,
        frame=(144.37, 461.63, 468.33, 690.99),
        axis_x=(0.0, 42.0),
        axis_y=(0.0, 1.0),
        x_units="in",
        y_units="dimensionless",
        curves=(
            CurveSpec("H_ft=6", (DrawingPart(107),)),
            CurveSpec("H_ft=10", (DrawingPart(106),)),
            CurveSpec("H_ft=14", (DrawingPart(105),)),
        ),
        uniform_count=17,
        graph_note="For LG >= 10 ft.",
    ),
}


# Viewer/export clips include the plotted graph, primary/secondary axis tick labels,
# and axis titles only. They intentionally exclude page headers, notes, captions, and
# surrounding body text.
GRAPH_CLIPS_BY_FIGURE: dict[str, tuple[float, float, float, float]] = {
    "03": (35.0, 84.0, 592.0, 364.0),
    "04": (82.0, 397.0, 532.0, 690.0),
    "05": (76.0, 86.0, 532.0, 386.0),
    "07": (88.0, 92.0, 486.0, 455.0),
    "08": (82.0, 74.0, 520.0, 336.0),
    "09": (86.0, 116.0, 520.0, 424.0),
    "10": (78.0, 92.0, 526.0, 386.0),
    "11": (82.0, 404.0, 540.0, 714.0),
    "12": (96.0, 98.0, 526.0, 382.0),
    "13": (76.0, 418.0, 526.0, 718.0),
    "14": (82.0, 162.0, 536.0, 434.0),
    "15": (90.0, 106.0, 516.0, 414.0),
    "16": (82.0, 78.0, 532.0, 354.0),
    "17": (82.0, 404.0, 532.0, 684.0),
    "18A": (76.0, 88.0, 526.0, 386.0),
    "18B": (74.0, 424.0, 520.0, 720.0),
}


def ticks(pairs: tuple[tuple[float, float], ...]) -> tuple[AxisTick, ...]:
    return tuple(AxisTick(float(value), float(page_coord), str(value)) for value, page_coord in pairs)


def tick_pairs(values: tuple[float, ...], page_coords: tuple[float, ...]) -> tuple[tuple[float, float], ...]:
    if len(values) != len(page_coords):
        raise ValueError("Tick values and page coordinates must have the same length.")
    return tuple(zip(values, page_coords))


def tick_values(start: float, step: float, count: int) -> tuple[float, ...]:
    return tuple(round(start + i * step, 10) for i in range(count))


def x_cal(source_label: str, pairs: tuple[tuple[float, float], ...], ignored: str = "none") -> PiecewiseAxisCalibration:
    return PiecewiseAxisCalibration("x", "page_x", source_label, ignored, ticks(pairs))


def y_cal(source_label: str, pairs: tuple[tuple[float, float], ...], ignored: str = "right-side duplicate or secondary ticks when present") -> PiecewiseAxisCalibration:
    return PiecewiseAxisCalibration("y", "page_y", source_label, ignored, ticks(pairs))


def depth_cal(source_label: str, pairs: tuple[tuple[float, float], ...]) -> PiecewiseAxisCalibration:
    return PiecewiseAxisCalibration("x", "page_y", source_label, "right metric depth axis", ticks(pairs))


def factor_x_cal(source_label: str, pairs: tuple[tuple[float, float], ...]) -> PiecewiseAxisCalibration:
    return PiecewiseAxisCalibration("y", "page_x", source_label, "none", ticks(pairs))


CALIBRATIONS_BY_FIGURE: dict[str, dict[str, PiecewiseAxisCalibration]] = {
    "Figure 3": {
        "x": x_cal(
            "bottom primary tw/D visible ticks",
            tick_pairs(
                tuple(i * 0.005 for i in range(17)),
                (112.45, 141.10, 169.26, 195.93, 224.09, 251.75, 278.92, 307.08, 334.74, 361.91, 389.57, 417.23, 444.40, 471.57, 499.23, 526.40, 551.59),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary KHe visible ticks",
            tick_pairs(
                tuple(float(i * 1000) for i in range(13)),
                (330.57, 303.89, 285.86, 267.09, 248.82, 230.29, 212.26, 193.74, 175.96, 157.43, 139.90, 120.63, 102.35),
            ),
        ),
    },
    "Figure 4": {
        "x": x_cal(
            "bottom primary H/Bd visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 17),
                (161.94, 182.85, 203.47, 223.82, 243.90, 263.98, 284.60, 305.23, 325.58, 345.93, 366.56, 386.63, 406.98, 427.34, 447.69, 468.31, 487.56),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary Be visible ticks",
            tick_pairs(
                tick_values(0.0, 0.1, 16),
                (665.24, 649.16, 633.07, 616.85, 600.48, 583.98, 567.48, 550.98, 534.89, 518.67, 502.58, 486.62, 470.81, 454.31, 438.36, 422.13),
            ),
        ),
    },
    "Figure 5": {
        "x": x_cal("bottom primary Bd/D labeled ticks", ((1.00, 155.79), (1.05, 208.06), (1.10, 262.44), (1.15, 317.94), (1.20, 373.44), (1.25, 427.94), (1.30, 479.21))),
        "y": y_cal("left primary Ee labeled ticks", ((0.8, 356.43), (0.9, 319.17), (1.0, 282.17), (1.1, 245.29), (1.2, 209.29), (1.3, 172.69), (1.4, 136.97)), "none"),
    },
    "Figure 7": {
        "x": depth_cal("left primary depth H feet labeled ticks", ((0.0, 128.27), (5.0, 181.07), (10.0, 233.32), (15.0, 286.12), (20.0, 340.02), (25.0, 392.83), (30.0, 447.28))),
        "y": factor_x_cal("top primary Fi labeled ticks", ((1.00, 195.69), (1.25, 249.05), (1.50, 302.40), (1.75, 354.10), (2.00, 406.35))),
    },
    "Figure 8": {
        "x": x_cal(
            "bottom primary tw/D visible ticks",
            tick_pairs(
                tick_values(0.0, 0.005, 17),
                (161.41, 182.56, 201.62, 220.94, 240.00, 260.63, 279.69, 299.53, 318.59, 339.22, 358.28, 377.34, 396.40, 417.54, 436.60, 456.71, 475.76),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary KHr visible ticks",
            tick_pairs(
                tick_values(0.0, 50.0, 11),
                (303.96, 283.34, 264.28, 244.17, 225.11, 206.05, 185.95, 166.37, 146.78, 127.20, 107.62),
            ),
        ),
    },
    "Figure 9": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 22),
                (171.59, 185.10, 199.64, 213.66, 228.21, 242.75, 256.78, 271.32, 285.86, 299.37, 313.91, 328.45, 343.00, 357.02, 371.57, 385.96, 400.13, 414.68, 429.22, 443.76, 458.30, 471.29),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal("left primary GHr labeled ticks", ((0.0, 366.03), (0.25, 322.40), (0.50, 279.29), (0.75, 236.70), (1.00, 194.62), (1.25, 151.51))),
    },
    "Figure 10": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 22),
                (163.51, 179.31, 194.61, 210.04, 224.57, 239.48, 254.40, 269.44, 284.74, 299.27, 314.06, 329.36, 344.15, 358.68, 373.85, 388.39, 403.94, 419.11, 433.51, 448.94, 463.86, 477.62),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal(
            "left primary NH visible ticks",
            tick_pairs(
                tick_values(0.5, 0.1, 16),
                (360.38, 346.62, 331.44, 315.25, 299.83, 284.15, 268.59, 252.40, 236.60, 220.53, 204.98, 189.17, 173.49, 157.81, 142.13, 127.85),
            ),
        ),
    },
    "Figure 11": {
        "x": x_cal(
            "bottom primary tw/D visible ticks",
            tick_pairs(
                tick_values(0.0, 0.005, 17),
                (167.15, 188.05, 208.54, 229.03, 249.52, 270.56, 291.05, 311.81, 332.30, 353.61, 374.10, 394.86, 415.35, 436.12, 456.61, 477.92, 497.72),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary KLr visible ticks",
            tick_pairs(
                tick_values(0.0, 50.0, 13),
                (680.37, 660.02, 639.53, 619.04, 598.56, 578.34, 557.58, 536.81, 515.78, 495.01, 474.25, 453.76, 433.41),
            ),
        ),
    },
    "Figure 12": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tuple(float(i) for i in range(0, 44, 2)),
                (180.20, 193.06, 207.42, 221.40, 235.39, 249.74, 264.35, 278.46, 292.81, 306.80, 321.15, 335.51, 349.99, 363.97, 378.33, 392.19, 406.54, 420.90, 435.63, 449.73, 464.09, 480.18),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal(
            "left primary GLr visible ticks",
            tick_pairs(
                tuple(i * 0.25 for i in range(11)),
                (347.99, 323.86, 304.80, 283.77, 262.36, 241.50, 219.79, 198.43, 176.97, 155.43, 132.66),
            ),
        ),
    },
    "Figure 13": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tuple(float(i) for i in range(0, 44, 2)),
                (158.98, 168.81, 183.05, 199.35, 214.62, 228.86, 243.87, 259.14, 273.37, 288.64, 303.91, 318.15, 333.16, 348.43, 362.66, 377.93, 393.20, 407.44, 422.45, 437.72, 451.95, 471.62),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal(
            "left primary NL visible ticks",
            tick_pairs(
                tick_values(0.5, 0.1, 16),
                (687.91, 672.26, 656.86, 640.94, 624.89, 609.50, 593.32, 577.92, 562.00, 545.96, 529.78, 514.77, 499.37, 483.46, 467.41, 446.18),
            ),
        ),
    },
    "Figure 14": {
        "x": x_cal(
            "bottom primary tw/D visible ticks",
            tick_pairs(
                tick_values(0.0, 0.005, 17),
                (175.96, 195.00, 214.79, 233.83, 254.12, 273.15, 292.69, 311.73, 331.52, 350.56, 370.85, 389.89, 409.68, 428.72, 448.51, 467.54, 488.08),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary KHh visible ticks",
            tick_pairs(
                tick_values(0.0, 2.5, 11),
                (377.81, 358.27, 339.23, 319.32, 300.28, 279.74, 260.70, 240.91, 221.87, 201.96, 182.42),
            ),
        ),
    },
    "Figure 15": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tuple(float(i) for i in range(0, 44, 2)),
                (158.93, 173.31, 187.69, 202.60, 217.50, 231.88, 246.79, 262.74, 277.12, 292.02, 306.93, 321.31, 336.22, 351.90, 366.29, 381.19, 396.62, 411.00, 425.91, 441.59, 455.98, 472.18),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal("left primary GHh labeled ticks", ((0.0, 369.40), (0.5, 315.02), (1.0, 259.06), (1.5, 203.37), (2.0, 146.10))),
    },
    "Figure 16": {
        "x": x_cal(
            "bottom primary tw/D visible ticks",
            tick_pairs(
                tick_values(0.0, 0.005, 17),
                (170.23, 189.47, 209.47, 228.71, 249.21, 268.45, 288.20, 307.43, 327.43, 346.67, 367.17, 386.41, 406.41, 425.65, 445.65, 464.89, 485.64),
            ),
            "top duplicate ticks",
        ),
        "y": y_cal(
            "left primary KLh visible ticks",
            tick_pairs(
                tick_values(0.0, 2.5, 11),
                (310.10, 290.36, 271.12, 251.00, 231.76, 211.00, 191.76, 171.76, 152.53, 132.40, 112.66),
            ),
        ),
    },
    "Figure 17": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 22),
                (166.19, 180.56, 194.93, 209.82, 224.71, 239.08, 253.98, 269.91, 284.28, 299.17, 314.07, 328.44, 343.33, 359.01, 373.38, 388.27, 403.68, 418.05, 432.95, 448.62, 462.99, 479.18),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal("left primary GLh labeled ticks", ((0.0, 657.73), (0.5, 621.15), (1.0, 583.53), (1.5, 545.91), (2.0, 508.29), (2.5, 470.67), (3.0, 434.62))),
    },
    "Figure 18-A": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 22),
                (154.29, 169.96, 184.81, 200.76, 216.44, 231.29, 247.24, 262.09, 276.94, 292.89, 308.29, 323.14, 339.10, 353.95, 368.80, 384.75, 400.15, 415.00, 430.95, 446.63, 462.30, 478.26),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal(
            "left primary RF visible ticks",
            tick_pairs(
                tick_values(0.0, 0.1, 11),
                (357.95, 335.54, 312.44, 289.06, 266.51, 243.00, 220.03, 196.93, 173.55, 151.00, 128.59),
            ),
        ),
    },
    "Figure 18-B": {
        "x": x_cal(
            "bottom primary diameter D inches visible ticks",
            tick_pairs(
                tick_values(0.0, 2.0, 22),
                (144.37, 160.03, 174.88, 190.84, 206.51, 221.36, 237.31, 252.16, 267.02, 282.97, 298.37, 313.22, 329.17, 344.02, 358.87, 374.82, 390.23, 405.08, 421.03, 436.70, 452.38, 468.33),
            ),
            "top metric diameter axis",
        ),
        "y": y_cal(
            "left primary RF visible ticks",
            tick_pairs(
                tick_values(0.0, 0.1, 11),
                (690.99, 668.58, 645.48, 622.10, 599.55, 576.04, 553.07, 529.97, 506.60, 484.05, 461.63),
            ),
        ),
    },
}


def axis_calibrations(spec: FigureSpec) -> dict[str, PiecewiseAxisCalibration]:
    return CALIBRATIONS_BY_FIGURE[spec.figure]


SPREADSHEET_SPECS: tuple[SpreadsheetSpec, ...] = (
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "03", "KHe", 17, 38, (TablePair(1, 2, "E_prime_ksi=0.2"), TablePair(3, 4, "E_prime_ksi=0.5"), TablePair(5, 6, "E_prime_ksi=1.0"), TablePair(7, 8, "E_prime_ksi=2.0"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "03", "KHe", 18, 39, (TablePair(1, 2, "E_prime_ksi=0.2"), TablePair(3, 4, "E_prime_ksi=0.5"), TablePair(5, 6, "E_prime_ksi=1.0"), TablePair(7, 8, "E_prime_ksi=2.0"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "04", "Be", 43, 76, (TablePair(1, 2, "soil_type=A"), TablePair(3, 4, "soil_type=B"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "04", "Be", 44, 77, (TablePair(1, 2, "soil_type=A"), TablePair(3, 4, "soil_type=B"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "05", "Ee", 80, 87, (TablePair(1, 2, "Ee"),)),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "05", "Ee", 81, 88, (TablePair(1, 2, "Ee"),)),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "07", "Fi", 91, 94, (TablePair(1, 2, "crossing=highway"),)),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "07", "Fi", 92, 95, (TablePair(1, 2, "crossing=railroad"),)),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "08", "KHr", 100, 111, (TablePair(1, 2, "Er_ksi=5"), TablePair(3, 4, "Er_ksi=10"), TablePair(5, 6, "Er_ksi=20"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "09", "GHr", 124, 135, (TablePair(1, 2, "H_ft=6"), TablePair(3, 4, "H_ft=10"), TablePair(5, 6, "H_ft=14"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "10", "NH", 140, 143, (TablePair(1, 2, "H_ft=6"), TablePair(3, 4, "H_ft=10"), TablePair(5, 6, "H_ft=14"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "11", "KLr", 246, 253, (TablePair(1, 2, "Er_ksi=5"), TablePair(3, 4, "Er_ksi=10"), TablePair(5, 6, "Er_ksi=20"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "12", "GLr", 262, 271, (TablePair(1, 2, "H_ft=6"), TablePair(3, 4, "H_ft=10"), TablePair(5, 6, "H_ft=14"))),
    SpreadsheetSpec("Railroad", RAILROAD_WORKBOOK, "13", "NL", 276, 281, (TablePair(1, 2, "H_ft=6"), TablePair(3, 4, "H_ft=10"), TablePair(5, 6, "H_ft=14"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "14", "KHh", 99, 118, (TablePair(1, 2, "Er_ksi=5"), TablePair(3, 4, "Er_ksi=10"), TablePair(5, 6, "Er_ksi=20"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "15", "GHh", 123, 138, (TablePair(1, 2, "H_ft=3_to_4"), TablePair(3, 4, "H_ft=6"), TablePair(5, 6, "H_ft=8"), TablePair(7, 8, "H_ft=10"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "16", "KLh", 244, 255, (TablePair(1, 2, "Er_ksi=5"), TablePair(3, 4, "Er_ksi=10"), TablePair(5, 6, "Er_ksi=20"))),
    SpreadsheetSpec("Highway", HIGHWAY_WORKBOOK, "17", "GLh", 260, 274, (TablePair(1, 2, "H_ft=3_to_4"), TablePair(3, 4, "H_ft=6"), TablePair(5, 6, "H_ft=8"), TablePair(7, 8, "H_ft=10"))),
)


ANNEX_CONTROLS: tuple[dict[str, Any], ...] = (
    {"example": "Annex B highway example", "figure": "03", "factor": "KHe", "curve_name": "E_prime_ksi=0.5", "x": 0.020, "api": 3024.0, "params": "tw/D=0.020; E_prime=0.5 ksi", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "04", "factor": "Be", "curve_name": "soil_type=A", "x": 4.9, "api": 1.09, "params": "H/Bd=4.9; soil_type=A", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "05", "factor": "Ee", "curve_name": "Ee", "x": 1.16, "api": 1.11, "params": "Bd/D=1.16", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "07", "factor": "Fi", "curve_name": "crossing=highway", "x": 6.0, "api": 1.47, "params": "H=6 ft; highway", "workbook": "Highway"},
    {"example": "Annex B railroad example", "figure": "08", "factor": "KHr", "curve_name": "Er_ksi=10", "x": 0.020, "api": 332.0, "params": "tw/D=0.020; Er=10 ksi", "workbook": "Railroad"},
    {"example": "Annex B railroad example", "figure": "09", "factor": "GHr", "curve_name": "H_ft=6", "x": 12.75, "api": 0.98, "params": "D=12.75 in; H=6 ft", "workbook": "Railroad"},
    {"example": "Annex B railroad example", "figure": "10", "factor": "NH", "curve_name": "H_ft=6", "x": 12.75, "api": 1.11, "params": "D=12.75 in; H=6 ft; double track", "workbook": "Railroad"},
    {"example": "Annex B railroad example", "figure": "11", "factor": "KLr", "curve_name": "Er_ksi=10", "x": 0.020, "api": 317.0, "params": "tw/D=0.020; Er=10 ksi", "workbook": "Railroad"},
    {"example": "Annex B railroad example", "figure": "12", "factor": "GLr", "curve_name": "H_ft=6", "x": 12.75, "api": 0.98, "params": "D=12.75 in; H=6 ft", "workbook": "Railroad"},
    {"example": "Annex B railroad example", "figure": "13", "factor": "NL", "curve_name": "H_ft=6", "x": 12.75, "api": 1.00, "params": "D=12.75 in; H=6 ft; double track", "workbook": "Railroad"},
    {"example": "Annex B highway example", "figure": "14", "factor": "KHh", "curve_name": "Er_ksi=10", "x": 0.020, "api": 14.3, "params": "tw/D=0.020; Er=10 ksi", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "15", "factor": "GHh", "curve_name": "H_ft=6", "x": 12.75, "api": 0.99, "params": "D=12.75 in; H=6 ft", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "16", "factor": "KLh", "curve_name": "Er_ksi=10", "x": 0.020, "api": 9.9, "params": "tw/D=0.020; Er=10 ksi", "workbook": "Highway"},
    {"example": "Annex B highway example", "figure": "17", "factor": "GLh", "curve_name": "H_ft=6", "x": 12.75, "api": 1.01, "params": "D=12.75 in; H=6 ft", "workbook": "Highway"},
)


def numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result) or math.isinf(result):
        return None
    return result


def source_page(spec: FigureSpec) -> str:
    return f"PDF page {spec.pdf_page} / API page {spec.api_page}"


def page_to_data(spec: FigureSpec, page_x: float, page_y: float) -> tuple[float, float]:
    calibrations = axis_calibrations(spec)
    if spec.orientation == "depth_on_y":
        return calibrations["x"].page_to_value(page_y), calibrations["y"].page_to_value(page_x)
    return calibrations["x"].page_to_value(page_x), calibrations["y"].page_to_value(page_y)


def data_to_page(spec: FigureSpec, x_value: float, y_value: float) -> tuple[float, float]:
    calibrations = axis_calibrations(spec)
    if spec.orientation == "depth_on_y":
        return calibrations["y"].value_to_page(y_value), calibrations["x"].value_to_page(x_value)
    return calibrations["x"].value_to_page(x_value), calibrations["y"].value_to_page(y_value)


def inside_frame(spec: FigureSpec, page_x: float, page_y: float, tolerance: float = 0.75) -> bool:
    fx0, fy0, fx1, fy1 = spec.frame
    return fx0 - tolerance <= page_x <= fx1 + tolerance and fy0 - tolerance <= page_y <= fy1 + tolerance


def line_points(p0: fitz.Point, p1: fitz.Point, steps: int = LINE_STEPS) -> list[tuple[float, float]]:
    return [
        (p0.x + (p1.x - p0.x) * i / steps, p0.y + (p1.y - p0.y) * i / steps)
        for i in range(steps + 1)
    ]


def cubic_points(points: tuple[fitz.Point, fitz.Point, fitz.Point, fitz.Point], steps: int = BEZIER_STEPS) -> list[tuple[float, float]]:
    p0, p1, p2, p3 = points
    out: list[tuple[float, float]] = []
    for i in range(steps + 1):
        t = i / steps
        mt = 1.0 - t
        x = mt**3 * p0.x + 3 * mt**2 * t * p1.x + 3 * mt * t**2 * p2.x + t**3 * p3.x
        y = mt**3 * p0.y + 3 * mt**2 * t * p1.y + 3 * mt * t**2 * p2.y + t**3 * p3.y
        out.append((x, y))
    return out


def sample_part(doc: fitz.Document, spec: FigureSpec, part: DrawingPart) -> list[DensePoint]:
    drawing = doc[spec.pdf_page - 1].get_drawings()[part.drawing_id]
    selected = set(part.item_indices) if part.item_indices is not None else None
    points: list[DensePoint] = []
    for item_index, item in enumerate(drawing["items"]):
        if selected is not None and item_index not in selected:
            continue
        item_type = item[0]
        if item_type == "l":
            raw_points = line_points(item[1], item[2])
        elif item_type == "c":
            raw_points = cubic_points((item[1], item[2], item[3], item[4]))
        else:
            continue
        for page_x, page_y in raw_points:
            if not inside_frame(spec, page_x, page_y):
                continue
            try:
                x_value, y_value = page_to_data(spec, page_x, page_y)
            except CalibrationRangeError:
                continue
            points.append(DensePoint(x_value, y_value, page_x, page_y))
    return points


def control_xs_for_part(doc: fitz.Document, spec: FigureSpec, part: DrawingPart) -> set[float]:
    drawing = doc[spec.pdf_page - 1].get_drawings()[part.drawing_id]
    selected = set(part.item_indices) if part.item_indices is not None else None
    xs: set[float] = set()
    for item_index, item in enumerate(drawing["items"]):
        if selected is not None and item_index not in selected:
            continue
        if item[0] == "l":
            endpoint_points = (item[1], item[2])
        elif item[0] == "c":
            endpoint_points = (item[1], item[4])
        else:
            continue
        for point in endpoint_points:
            if inside_frame(spec, point.x, point.y, tolerance=1.25):
                try:
                    xs.add(page_to_data(spec, point.x, point.y)[0])
                except CalibrationRangeError:
                    continue
    return xs


def dedupe_points(points: list[DensePoint], tolerance: float = 1e-7) -> list[DensePoint]:
    if not points:
        return []
    ordered = sorted(points, key=lambda p: (p.x, p.page_x, p.page_y))
    deduped: list[DensePoint] = [ordered[0]]
    for point in ordered[1:]:
        last = deduped[-1]
        if abs(point.x - last.x) <= tolerance:
            if abs(point.y - last.y) > abs(point.y - deduped[-1].y):
                deduped[-1] = point
            continue
        deduped.append(point)
    return deduped


def sample_curve(doc: fitz.Document, spec: FigureSpec, curve: CurveSpec) -> tuple[list[DensePoint], set[float]]:
    main_points: list[DensePoint] = []
    control_xs: set[float] = set()
    for part in curve.parts:
        main_points.extend(sample_part(doc, spec, part))
        control_xs.update(control_xs_for_part(doc, spec, part))
    main_points = dedupe_points(main_points)
    if curve.tail_drawing_id is not None and main_points:
        merge_x = max(p.x for p in main_points)
        tail_points = sample_part(doc, spec, DrawingPart(curve.tail_drawing_id))
        tail_points = [p for p in tail_points if p.x >= merge_x - 1e-7]
        main_points.extend(tail_points)
        control_xs.update(x for x in control_xs_for_part(doc, spec, DrawingPart(curve.tail_drawing_id)) if x >= merge_x - 1e-7)
    return dedupe_points(main_points), control_xs


def interpolate_dense(points: list[DensePoint], x_value: float, *, tolerance: float = 1e-8) -> DensePoint | None:
    ordered = dedupe_points(points)
    if len(ordered) < 2:
        return None
    xmin = ordered[0].x
    xmax = ordered[-1].x
    if x_value < xmin - tolerance or x_value > xmax + tolerance:
        return None
    if abs(x_value - xmin) <= tolerance:
        return ordered[0]
    if abs(x_value - xmax) <= tolerance:
        return ordered[-1]
    for p0, p1 in zip(ordered, ordered[1:]):
        if p0.x - tolerance <= x_value <= p1.x + tolerance:
            if abs(p1.x - p0.x) <= tolerance:
                return p0
            fraction = (x_value - p0.x) / (p1.x - p0.x)
            return DensePoint(
                x=x_value,
                y=p0.y + fraction * (p1.y - p0.y),
                page_x=p0.page_x + fraction * (p1.page_x - p0.page_x),
                page_y=p0.page_y + fraction * (p1.page_y - p0.page_y),
            )
    return None


def round_for_output(value: float, digits: int = 6) -> float:
    rounded = round(float(value), digits)
    if rounded == 0:
        return 0.0
    return rounded


def read_spreadsheet_points() -> tuple[list[dict[str, Any]], dict[tuple[str, str], set[float]], dict[tuple[str, str, str, str], dict[float, float]]]:
    rows: list[dict[str, Any]] = []
    x_values_by_curve: dict[tuple[str, str], set[float]] = defaultdict(set)
    tables: dict[tuple[str, str, str, str], dict[float, float]] = defaultdict(dict)
    workbook_cache: dict[Path, openpyxl.Workbook] = {}
    for spec in SPREADSHEET_SPECS:
        workbook = workbook_cache.get(spec.path)
        if workbook is None:
            workbook = openpyxl.load_workbook(spec.path, data_only=True, read_only=True)
            workbook_cache[spec.path] = workbook
        sheet = workbook["Tables"]
        for pair in spec.pairs:
            for row_index in range(spec.row_start, spec.row_end + 1):
                x_value = numeric(sheet.cell(row_index, pair.x_col).value)
                y_value = numeric(sheet.cell(row_index, pair.y_col).value)
                if x_value is None or y_value is None:
                    continue
                row = {
                    "workbook": spec.workbook_label,
                    "figure": spec.figure,
                    "factor": spec.factor,
                    "curve_name": pair.curve_name,
                    "x_value": x_value,
                    "spreadsheet_value": y_value,
                    "cell": f"{get_column_letter(pair.y_col)}{row_index}",
                }
                rows.append(row)
                x_values_by_curve[(spec.figure, pair.curve_name)].add(x_value)
                tables[(spec.workbook_label, spec.figure, spec.factor, pair.curve_name)][x_value] = y_value
    return rows, x_values_by_curve, tables


def point_type_note(types: set[str]) -> str:
    order = ["endpoint/control", "spreadsheet_x", "annex_control_x", "uniform_sample"]
    return ";".join([name for name in order if name in types] + sorted(types - set(order)))


def generate_digitized_records(
    doc: fitz.Document,
    spreadsheet_xs: dict[tuple[str, str], set[float]],
) -> tuple[dict[str, list[DigitizedRecord]], dict[tuple[str, str], list[DensePoint]], dict[str, dict[str, Any]]]:
    records_by_figure: dict[str, list[DigitizedRecord]] = defaultdict(list)
    dense_by_curve: dict[tuple[str, str], list[DensePoint]] = {}
    metadata: dict[str, dict[str, Any]] = {}
    annex_xs: dict[tuple[str, str], set[float]] = defaultdict(set)
    for control in ANNEX_CONTROLS:
        annex_xs[(control["figure"], control["curve_name"])].add(float(control["x"]))

    for fig_id, spec in FIGURES.items():
        figure_meta: dict[str, Any] = {
            "figure": spec.figure,
            "factor": spec.factor,
            "pdf_page": spec.pdf_page,
            "api_page": spec.api_page,
            "source_page": source_page(spec),
            "frame_pdf_points": spec.frame,
            "axis_x_range": spec.axis_x,
            "axis_y_range": spec.axis_y,
            "calibration_supersedes": "first-pass endpoint-span affine calibration",
            "axis_calibrations": {
                axis_name: calibration.to_metadata()
                for axis_name, calibration in axis_calibrations(spec).items()
            },
            "orientation": spec.orientation,
            "x_units": spec.x_units,
            "y_units": spec.y_units,
            "graph_note": spec.graph_note,
            "curves": {},
        }
        for curve in spec.curves:
            dense_points, control_xs = sample_curve(doc, spec, curve)
            if len(dense_points) < 2:
                raise RuntimeError(f"{spec.figure} {curve.curve_name} produced too few points.")
            dense_by_curve[(fig_id, curve.curve_name)] = dense_points
            xmin = min(p.x for p in dense_points)
            xmax = max(p.x for p in dense_points)
            target_types: dict[float, set[str]] = defaultdict(set)
            for x_value in control_xs:
                if xmin - 1e-8 <= x_value <= xmax + 1e-8:
                    target_types[round_for_output(x_value, 8)].add("endpoint/control")
            for x_value in spreadsheet_xs.get((fig_id, curve.curve_name), set()):
                if xmin - 1e-8 <= x_value <= xmax + 1e-8:
                    target_types[round_for_output(x_value, 8)].add("spreadsheet_x")
            for x_value in annex_xs.get((fig_id, curve.curve_name), set()):
                if xmin - 1e-8 <= x_value <= xmax + 1e-8:
                    target_types[round_for_output(x_value, 8)].add("annex_control_x")
            if spec.uniform_count > 1:
                for i in range(spec.uniform_count):
                    x_value = xmin + (xmax - xmin) * i / (spec.uniform_count - 1)
                    target_types[round_for_output(x_value, 8)].add("uniform_sample")

            for rounded_x, types in sorted(target_types.items()):
                point = interpolate_dense(dense_points, rounded_x)
                if point is None:
                    continue
                notes = [f"point_type={point_type_note(types)}"]
                if curve.notes:
                    notes.append(curve.notes)
                if spec.graph_note:
                    notes.append(spec.graph_note)
                records_by_figure[fig_id].append(
                    DigitizedRecord(
                        figure=spec.figure,
                        factor=spec.factor,
                        curve_name=curve.curve_name,
                        x_value=round_for_output(point.x),
                        x_units=spec.x_units,
                        y_value=round_for_output(point.y),
                        y_units=spec.y_units,
                        source_page=source_page(spec),
                        digitization_method=DIGITIZATION_METHOD,
                        notes=" | ".join(notes),
                        point_type=point_type_note(types),
                        page_x=point.page_x,
                        page_y=point.page_y,
                    )
                )
            figure_meta["curves"][curve.curve_name] = {
                "drawing_parts": [
                    {"drawing_id": part.drawing_id, "item_indices": list(part.item_indices) if part.item_indices is not None else None}
                    for part in curve.parts
                ],
                "tail_drawing_id": curve.tail_drawing_id,
                "curve_x_range": [round_for_output(xmin), round_for_output(xmax)],
                "curve_y_range": [round_for_output(min(p.y for p in dense_points)), round_for_output(max(p.y for p in dense_points))],
                "notes": curve.notes,
            }
        metadata[fig_id] = figure_meta
    return records_by_figure, dense_by_curve, metadata


def write_raw_csvs(records_by_figure: dict[str, list[DigitizedRecord]]) -> None:
    fieldnames = [
        "figure",
        "factor",
        "curve_name",
        "x_value",
        "x_units",
        "y_value",
        "y_units",
        "source_page",
        "digitization_method",
        "notes",
    ]
    for fig_id, spec in FIGURES.items():
        path = OUTPUT_DIR / spec.filename
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for record in sorted(records_by_figure[fig_id], key=lambda r: (r.curve_name, r.x_value)):
                writer.writerow({name: getattr(record, name) for name in fieldnames})


def write_provenance(records_by_figure: dict[str, list[DigitizedRecord]], metadata: dict[str, dict[str, Any]], doc: fitz.Document) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OVERLAY_DIR.mkdir(parents=True, exist_ok=True)
    PAGE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    with (OUTPUT_DIR / "calibration_metadata.json").open("w", encoding="utf-8") as handle:
        json.dump(metadata, handle, indent=2)
    with (OUTPUT_DIR / "digitized_points_provenance.csv").open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "figure",
            "factor",
            "curve_name",
            "x_value",
            "y_value",
            "point_type",
            "page_x_pdf_points",
            "page_y_pdf_points",
            "source_page",
            "notes",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for fig_id in FIGURES:
            for record in sorted(records_by_figure[fig_id], key=lambda r: (r.curve_name, r.x_value)):
                writer.writerow(
                    {
                        "figure": record.figure,
                        "factor": record.factor,
                        "curve_name": record.curve_name,
                        "x_value": record.x_value,
                        "y_value": record.y_value,
                        "point_type": record.point_type,
                        "page_x_pdf_points": round_for_output(record.page_x),
                        "page_y_pdf_points": round_for_output(record.page_y),
                        "source_page": record.source_page,
                        "notes": record.notes,
                    }
                )
    rendered_pages = sorted({spec.pdf_page for spec in FIGURES.values()})
    for pdf_page in rendered_pages:
        page = doc[pdf_page - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False, annots=False)
        pix.save(PAGE_IMAGE_DIR / f"pdf_page_{pdf_page:02d}.png")


def graph_clip(spec: FigureSpec) -> fitz.Rect:
    fig_id = next((key for key, figure in FIGURES.items() if figure is spec), None)
    if fig_id is None:
        raise KeyError(f"Figure spec is not registered: {spec.figure}")
    x0, y0, x1, y1 = GRAPH_CLIPS_BY_FIGURE[fig_id]
    return fitz.Rect(max(0, x0), max(0, y0), min(612, x1), min(792, y1))


def write_overlays(doc: fitz.Document, records_by_figure: dict[str, list[DigitizedRecord]]) -> None:
    colors = [
        (214, 39, 40),
        (31, 119, 180),
        (44, 160, 44),
        (148, 103, 189),
        (255, 127, 14),
        (23, 190, 207),
    ]
    scale = 3
    for fig_id, spec in FIGURES.items():
        clip = graph_clip(spec)
        page = doc[spec.pdf_page - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip, alpha=False, annots=False)
        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        draw = ImageDraw.Draw(image)
        curve_order = {curve.curve_name: index for index, curve in enumerate(spec.curves)}
        calibrations = axis_calibrations(spec)
        for record in records_by_figure[fig_id]:
            color = colors[curve_order.get(record.curve_name, 0) % len(colors)]
            cx = (record.page_x - clip.x0) * scale
            cy = (record.page_y - clip.y0) * scale
            radius = 3 if "spreadsheet_x" not in record.point_type else 4
            draw.ellipse((cx - radius, cy - radius, cx + radius, cy + radius), outline=color, fill=color)
        tick_color = (0, 150, 170)
        fx0, fy0, fx1, fy1 = spec.frame
        for tick in calibrations["x"].ticks:
            page_x, page_y = (fx0, tick.page_coord) if spec.orientation == "depth_on_y" else (tick.page_coord, fy1)
            cx = (page_x - clip.x0) * scale
            cy = (page_y - clip.y0) * scale
            size = 5
            draw.line((cx - size, cy, cx + size, cy), fill=tick_color, width=2)
            draw.line((cx, cy - size, cx, cy + size), fill=tick_color, width=2)
        for tick in calibrations["y"].ticks:
            page_x, page_y = (tick.page_coord, fy0) if spec.orientation == "depth_on_y" else (fx0, tick.page_coord)
            cx = (page_x - clip.x0) * scale
            cy = (page_y - clip.y0) * scale
            size = 5
            draw.rectangle((cx - size, cy - size, cx + size, cy + size), outline=tick_color, width=2)
        for curve in spec.curves:
            curve_records = [r for r in records_by_figure[fig_id] if r.curve_name == curve.curve_name]
            if not curve_records:
                continue
            label_record = max(curve_records, key=lambda r: r.x_value)
            color = colors[curve_order[curve.curve_name] % len(colors)]
            tx = (label_record.page_x - clip.x0) * scale + 8
            ty = (label_record.page_y - clip.y0) * scale - 8
            draw.text((tx, ty), curve.curve_name, fill=color)
        overlay_name = spec.filename.replace(".csv", "_overlay.png")
        image.save(OVERLAY_DIR / overlay_name)


def write_viewer_assets(
    doc: fitz.Document,
    records_by_figure: dict[str, list[DigitizedRecord]],
    metadata: dict[str, dict[str, Any]],
) -> None:
    CLEAN_GRAPH_DIR.mkdir(parents=True, exist_ok=True)
    scale = 3
    figures: list[dict[str, Any]] = []
    for fig_id, spec in FIGURES.items():
        clip = graph_clip(spec)
        page = doc[spec.pdf_page - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip, alpha=False, annots=False)
        underlay_name = spec.filename.replace(".csv", "_underlay.png")
        underlay_path = CLEAN_GRAPH_DIR / underlay_name
        pix.save(underlay_path)
        overlay_name = spec.filename.replace(".csv", "_overlay.png")
        figure_records = records_by_figure[fig_id]
        points_by_curve = Counter(record.curve_name for record in figure_records)
        figures.append(
            {
                "id": fig_id,
                "figure": spec.figure,
                "factor": spec.factor,
                "filename": spec.filename,
                "source_page": source_page(spec),
                "pdf_page": spec.pdf_page,
                "api_page": spec.api_page,
                "x_units": spec.x_units,
                "y_units": spec.y_units,
                "orientation": spec.orientation,
                "frame_pdf_points": list(spec.frame),
                "axis_x_range": list(spec.axis_x),
                "axis_y_range": list(spec.axis_y),
                "clip_pdf_points": [round_for_output(clip.x0), round_for_output(clip.y0), round_for_output(clip.x1), round_for_output(clip.y1)],
                "render_scale": scale,
                "image_size_px": [pix.width, pix.height],
                "underlay_url": f"/api/digitized-assets/graph_underlays/{underlay_name}",
                "overlay_url": f"/api/digitized-assets/overlays/{overlay_name}",
                "csv_url": f"/api/digitized-assets/{spec.filename}",
                "calibrations": metadata[fig_id]["axis_calibrations"],
                "curves": [
                    {
                        "curve_name": curve.curve_name,
                        "point_count": points_by_curve[curve.curve_name],
                        "x_range": metadata[fig_id]["curves"][curve.curve_name]["curve_x_range"],
                        "y_range": metadata[fig_id]["curves"][curve.curve_name]["curve_y_range"],
                        "notes": curve.notes,
                    }
                    for curve in spec.curves
                ],
            }
        )
    manifest = {
        "schema_version": 1,
        "source": "API RP 1102 digitized graph package",
        "digitization_method": DIGITIZATION_METHOD,
        "figures": figures,
    }
    with (OUTPUT_DIR / "viewer_manifest.json").open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2)


def status_for_difference(factor: str, digitized: float | None, spreadsheet: float | None) -> tuple[str, float | None, float | None, str]:
    if digitized is None or spreadsheet is None:
        return "REVIEW", None, None, "Spreadsheet point is outside the digitized curve range or is nonnumeric; treated as guardrail/non-API-derived."
    abs_diff = abs(spreadsheet - digitized)
    pct_diff = None if digitized == 0 else abs_diff / abs(digitized) * 100.0
    if factor == "Fi":
        if abs_diff <= 0.01:
            return "PASS", abs_diff, pct_diff, ""
        if abs_diff <= 0.03:
            return "REVIEW", abs_diff, pct_diff, ""
        return "FAIL", abs_diff, pct_diff, ""
    if factor in {"KHe", "KHr", "KLr", "KHh", "KLh"}:
        if pct_diff is not None and pct_diff <= 1.5:
            return "PASS", abs_diff, pct_diff, ""
        if pct_diff is not None and pct_diff <= 3.0:
            return "REVIEW", abs_diff, pct_diff, ""
        return "FAIL", abs_diff, pct_diff, ""
    if abs_diff <= 0.02 or (pct_diff is not None and pct_diff <= 2.0):
        return "PASS", abs_diff, pct_diff, ""
    if pct_diff is not None and pct_diff <= 4.0:
        return "REVIEW", abs_diff, pct_diff, ""
    return "FAIL", abs_diff, pct_diff, ""


def digitized_value_for(dense_by_curve: dict[tuple[str, str], list[DensePoint]], figure: str, curve_name: str, x_value: float) -> float | None:
    point = interpolate_dense(dense_by_curve[(figure, curve_name)], x_value)
    return None if point is None else point.y


def write_spreadsheet_comparison(
    spreadsheet_rows: list[dict[str, Any]],
    dense_by_curve: dict[tuple[str, str], list[DensePoint]],
) -> list[dict[str, Any]]:
    comparison_rows: list[dict[str, Any]] = []
    for row in spreadsheet_rows:
        digitized_value = digitized_value_for(dense_by_curve, row["figure"], row["curve_name"], row["x_value"])
        status, abs_diff, pct_diff, note = status_for_difference(row["factor"], digitized_value, row["spreadsheet_value"])
        if row["x_value"] > FIGURES[row["figure"]].axis_x[1] and FIGURES[row["figure"]].orientation != "depth_on_y":
            note = (note + " " if note else "") + "Spreadsheet x-value is beyond graph axis maximum."
        comparison_rows.append(
            {
                "workbook": row["workbook"],
                "figure": FIGURES[row["figure"]].figure,
                "factor": row["factor"],
                "curve_name": row["curve_name"],
                "x_value": row["x_value"],
                "digitized_api_value": "" if digitized_value is None else round_for_output(digitized_value),
                "spreadsheet_value": row["spreadsheet_value"],
                "absolute_difference": "" if abs_diff is None else round_for_output(abs_diff),
                "percent_difference": "" if pct_diff is None else round_for_output(pct_diff),
                "status": status,
                "notes": f"{note} spreadsheet_cell=Tables!{row['cell']}".strip(),
            }
        )
    path = OUTPUT_DIR / "spreadsheet_table_comparison.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "workbook",
            "figure",
            "factor",
            "curve_name",
            "x_value",
            "digitized_api_value",
            "spreadsheet_value",
            "absolute_difference",
            "percent_difference",
            "status",
            "notes",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(comparison_rows)
    return comparison_rows


def spreadsheet_table_value(
    workbook_tables: dict[tuple[str, str, str, str], dict[float, float]],
    workbook: str,
    figure: str,
    factor: str,
    curve_name: str,
    x_value: float,
) -> float | None:
    table = workbook_tables.get((workbook, figure, factor, curve_name))
    if not table:
        return None
    try:
        return linear_interpolate(f"{workbook} {figure} {curve_name}", table, x_value, on_out_of_range="warn").value
    except DigitizedRangeError:
        return None


def app_table_value(factor: str, curve_name: str, x_value: float, workbook: str) -> float | None:
    try:
        if workbook == "Highway":
            import app.standards.highway_tables as tables
        else:
            import app.standards.railroad_tables as tables
    except Exception:
        return None

    table: dict[float, float] | None = None
    if factor == "KHe":
        key = {"E_prime_ksi=0.2": 200.0, "E_prime_ksi=0.5": 500.0, "E_prime_ksi=1.0": 1000.0, "E_prime_ksi=2.0": 2000.0}[curve_name]
        table = tables.EARTH_KHE_BY_E_PRIME[key]
    elif factor == "Be":
        table = tables.BURIAL_A_BY_H_BD if curve_name == "soil_type=A" else tables.BURIAL_B_BY_H_BD
    elif factor == "Ee":
        table = tables.EXCAVATION_BY_BD_D
    elif factor == "Fi":
        table = tables.IMPACT_BY_COVER
    elif factor in {"KHr", "KHh"}:
        key = {"Er_ksi=5": 5000.0, "Er_ksi=10": 10000.0, "Er_ksi=20": 20000.0}[curve_name]
        table = tables.KH_BY_ER[key]
    elif factor in {"KLr", "KLh"}:
        key = {"Er_ksi=5": 5000.0, "Er_ksi=10": 10000.0, "Er_ksi=20": 20000.0}[curve_name]
        table = tables.KL_BY_ER[key]
    elif factor in {"GHr", "GHh"}:
        depth_key = curve_name.replace("H_ft=", "").replace("_to_4", "")
        table = tables.GH_BY_DEPTH[depth_key]
    elif factor in {"GLr", "GLh"}:
        depth_key = curve_name.replace("H_ft=", "").replace("_to_4", "")
        table = tables.GL_BY_DEPTH[depth_key]
    elif factor == "NH":
        depth_key = curve_name.replace("H_ft=", "")
        table = tables.NH_BY_DEPTH[depth_key]
    elif factor == "NL":
        depth_key = curve_name.replace("H_ft=", "")
        table = tables.NL_BY_DEPTH[depth_key]
    if table is None:
        return None
    try:
        return linear_interpolate(f"app {factor} {curve_name}", table, x_value, on_out_of_range="warn").value
    except Exception:
        return None


def write_annex_controls(
    dense_by_curve: dict[tuple[str, str], list[DensePoint]],
    workbook_tables: dict[tuple[str, str, str, str], dict[float, float]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for control in ANNEX_CONTROLS:
        digitized = digitized_value_for(dense_by_curve, control["figure"], control["curve_name"], control["x"])
        spreadsheet = spreadsheet_table_value(
            workbook_tables,
            control["workbook"],
            control["figure"],
            control["factor"],
            control["curve_name"],
            control["x"],
        )
        app_value = app_table_value(control["factor"], control["curve_name"], control["x"], control["workbook"])
        status, _, _, note = status_for_difference(control["factor"], digitized, control["api"])
        row = {
            "example": control["example"],
            "figure": FIGURES[control["figure"]].figure,
            "factor": control["factor"],
            "input_parameters": control["params"],
            "api_annex_b_value": control["api"],
            "digitized_value": "" if digitized is None else round_for_output(digitized),
            "spreadsheet_value": "" if spreadsheet is None else round_for_output(spreadsheet),
            "app_table_value": "" if app_value is None else round_for_output(app_value),
            "digitized_difference": "" if digitized is None else round_for_output(digitized - control["api"]),
            "spreadsheet_difference": "" if spreadsheet is None else round_for_output(spreadsheet - control["api"]),
            "app_table_difference": "" if app_value is None else round_for_output(app_value - control["api"]),
            "status": status,
            "notes": note,
        }
        rows.append(row)
    with (OUTPUT_DIR / "annex_b_control_point_check.csv").open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "example",
            "figure",
            "factor",
            "input_parameters",
            "api_annex_b_value",
            "digitized_value",
            "spreadsheet_value",
            "app_table_value",
            "digitized_difference",
            "spreadsheet_difference",
            "app_table_difference",
            "status",
            "notes",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return rows


def write_lookup_workbook(records_by_figure: dict[str, list[DigitizedRecord]]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for fig_id, spec in FIGURES.items():
        sheet = workbook.create_sheet(spec.sheet_name)
        headers = [
            "figure",
            "factor",
            "curve_name",
            "x_value",
            "x_units",
            "y_value",
            "y_units",
            "point_type",
            "source_page",
            "digitization_method",
            "notes",
        ]
        sheet.append(headers)
        for record in sorted(records_by_figure[fig_id], key=lambda r: (r.curve_name, r.x_value)):
            sheet.append(
                [
                    record.figure,
                    record.factor,
                    record.curve_name,
                    record.x_value,
                    record.x_units,
                    record.y_value,
                    record.y_units,
                    record.point_type,
                    record.source_page,
                    record.digitization_method,
                    record.notes,
                ]
            )
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [14, 10, 20, 12, 12, 12, 14, 28, 24, 48, 72]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
        for row in sheet.iter_rows(min_row=2, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet.max_row > 1:
            for row in sheet.iter_rows(min_row=2, min_col=4, max_col=6):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000000"
    workbook.save(OUTPUT_DIR / "api1102_digitized_lookup_tables.xlsx")


def summarize_curve_counts(records_by_figure: dict[str, list[DigitizedRecord]]) -> dict[str, Any]:
    curve_counts: dict[str, Any] = {}
    for fig_id, spec in FIGURES.items():
        counts = Counter(record.curve_name for record in records_by_figure[fig_id])
        curve_counts[fig_id] = {
            "figure": spec.figure,
            "factor": spec.factor,
            "curve_count": len(counts),
            "point_count": sum(counts.values()),
            "points_by_curve": dict(counts),
        }
    return curve_counts


def write_report(
    records_by_figure: dict[str, list[DigitizedRecord]],
    metadata: dict[str, dict[str, Any]],
    comparison_rows: list[dict[str, Any]],
    annex_rows: list[dict[str, Any]],
) -> None:
    curve_counts = summarize_curve_counts(records_by_figure)
    comparison_status = Counter(row["status"] for row in comparison_rows)
    annex_status = Counter(row["status"] for row in annex_rows)
    factor_status = defaultdict(Counter)
    for row in comparison_rows:
        factor_status[row["factor"]][row["status"]] += 1
    failing_factors = sorted(factor for factor, counts in factor_status.items() if counts["FAIL"])
    review_factors = sorted(factor for factor, counts in factor_status.items() if counts["REVIEW"] and not counts["FAIL"])
    pass_factors = sorted(factor for factor, counts in factor_status.items() if counts["PASS"] and not counts["REVIEW"] and not counts["FAIL"])

    lines: list[str] = []
    lines.append("# API RP 1102 Digitization Verification Report")
    lines.append("")
    lines.append("## Scope")
    lines.append("")
    lines.append("Digitized API RP 1102, Seventh Edition graph-derived factors for Figures 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18-A, and 18-B. No app calculation logic or Excel workbook logic was updated.")
    lines.append("")
    lines.append("## Method")
    lines.append("")
    lines.append("- Source PDF graphs were vector drawings, not low-resolution raster scans.")
    lines.append("- Curves were extracted from PDF drawing paths using PyMuPDF, including Bezier and line path commands.")
    lines.append("- Plot axes were calibrated with piecewise linear transforms from labeled API/US tick marks.")
    lines.append("- This package supersedes the earlier first-pass endpoint-span affine calibration.")
    lines.append("- Points were generated at curve endpoints/control breakpoints, existing spreadsheet x-values inside the curve range, Annex B control x-values, and uniform in-range samples.")
    lines.append("- OCR was not used for curve digitization.")
    lines.append("- The normalized lookup workbook was generated with `openpyxl` because `@oai/artifact-tool` was unavailable in this session.")
    lines.append("")
    lines.append("## Calibration")
    lines.append("")
    lines.append("| Figure | Factor | Source page | Axis x range | Axis y range | Output x units | Output y units | Curves | Points |")
    lines.append("|---|---:|---|---:|---:|---|---|---:|---:|")
    for fig_id, spec in FIGURES.items():
        counts = curve_counts[fig_id]
        lines.append(
            f"| {spec.figure} | {spec.factor} | {source_page(spec)} | {spec.axis_x[0]} to {spec.axis_x[1]} | {spec.axis_y[0]} to {spec.axis_y[1]} | {spec.x_units} | {spec.y_units} | {counts['curve_count']} | {counts['point_count']} |"
        )
    lines.append("")
    lines.append("Calibration metadata with PDF frame coordinates, labeled tick controls, and drawing IDs is saved in `calibration_metadata.json`; point-level PDF coordinates are saved in `digitized_points_provenance.csv`.")
    lines.append("")
    lines.append("## Calibration QA")
    lines.append("")
    lines.append("| Figure | x-axis tick controls | y-axis tick controls | Max page residual | Ignored axes/ticks |")
    lines.append("|---|---|---|---:|---|")
    for fig_id, spec in FIGURES.items():
        calibrations = axis_calibrations(spec)
        x_ticks = ", ".join(tick.label for tick in calibrations["x"].ticks)
        y_ticks = ", ".join(tick.label for tick in calibrations["y"].ticks)
        max_page_residual = max(
            calibrations["x"].residuals()["max_page_coord_residual"],
            calibrations["y"].residuals()["max_page_coord_residual"],
        )
        ignored = "; ".join(
            sorted(
                {
                    calibrations["x"].ignored_secondary_axis,
                    calibrations["y"].ignored_secondary_axis,
                }
                - {"none"}
            )
        )
        lines.append(f"| {spec.figure} | {x_ticks} | {y_ticks} | {max_page_residual:.6g} | {ignored or 'none'} |")
    lines.append("")
    lines.append("## Curves Captured")
    lines.append("")
    for fig_id, spec in FIGURES.items():
        counts = curve_counts[fig_id]["points_by_curve"]
        curve_text = ", ".join(f"{name} ({count} pts)" for name, count in counts.items())
        lines.append(f"- {spec.figure} {spec.factor}: {curve_text}")
    lines.append("")
    lines.append("## Assumptions And Limitations")
    lines.append("")
    lines.append("- The PDF vector paths are treated as the controlling representation of the published graphs.")
    lines.append("- The prior endpoint-span calibration has been superseded; use only the piecewise labeled-tick outputs in this folder.")
    lines.append("- Leader arrows, tick marks, labels, and diagram callouts were excluded by drawing ID and stroke geometry.")
    lines.append("- Figure 3 has a single visible common tail after KHe curves converge; that tail is reused for the higher E' curves after their unique strokes end and is identified in notes.")
    lines.append("- Spreadsheet rows outside the drawn/API graph range are not extrapolated; they are labeled as guardrail/non-API-derived review rows.")
    lines.append("- Figure 7 is plotted with Fi on the horizontal axis and depth on the vertical axis; the output tables normalize x to depth H for interpolation.")
    lines.append("- Digitized values are graph-derived and may differ from Annex B rounded example values because Annex B values are rounded to engineering precision.")
    lines.append("")
    lines.append("## Annex B Control Points")
    lines.append("")
    lines.append(f"Status counts: {dict(annex_status)}")
    lines.append("")
    lines.append("| Factor | Figure | API Annex B | Digitized | Spreadsheet | App table | Status |")
    lines.append("|---|---|---:|---:|---:|---:|---|")
    for row in annex_rows:
        lines.append(
            f"| {row['factor']} | {row['figure']} | {row['api_annex_b_value']} | {row['digitized_value']} | {row['spreadsheet_value']} | {row['app_table_value']} | {row['status']} |"
        )
    lines.append("")
    lines.append("## Spreadsheet Comparison")
    lines.append("")
    lines.append(f"Status counts: {dict(comparison_status)}")
    lines.append(f"Superseded first-pass endpoint-span status counts: {FIRST_PASS_ENDPOINT_SPAN_SUMMARY['comparison_status']}")
    if factor_status["KHe"]["FAIL"]:
        lines.append("KHe failures remain after tick-calibrated correction.")
    else:
        lines.append("KHe failures did not remain after tick-calibrated correction.")
    lines.append("")
    lines.append("| Factor | PASS | REVIEW | FAIL |")
    lines.append("|---|---:|---:|---:|")
    for factor in sorted(factor_status):
        counts = factor_status[factor]
        lines.append(f"| {factor} | {counts['PASS']} | {counts['REVIEW']} | {counts['FAIL']} |")
    lines.append("")
    lines.append("### Factors That Pass")
    lines.append("")
    lines.append(", ".join(pass_factors) if pass_factors else "None as a complete factor family.")
    lines.append("")
    lines.append("### Factors Needing Review")
    lines.append("")
    lines.append(", ".join(review_factors) if review_factors else "None as a complete factor family.")
    lines.append("")
    lines.append("### Factors That Fail")
    lines.append("")
    lines.append(", ".join(failing_factors) if failing_factors else "None.")
    lines.append("")
    lines.append("## Recommended Spreadsheet Corrections")
    lines.append("")
    if failing_factors:
        lines.append("- Replace failed factor table values with graph-derived values from `api1102_digitized_lookup_tables.xlsx`, prioritizing factors with Annex B mismatches and high FAIL counts.")
    else:
        lines.append("- No material value replacements were identified by the configured tolerances.")
    lines.append("- Remove, clamp, or explicitly label rows outside the API graph range as spreadsheet guardrails rather than API-derived values.")
    lines.append("- Update app standards tables only after the workbook corrections are reviewed and approved, because the current app tables mirror the workbook tables.")
    lines.append("- Preserve `calibration_metadata.json`, provenance CSV, and overlay PNGs with any future table revision so reviewers can reproduce the source values.")
    lines.append("")
    lines.append("## Source Of Truth Conclusion")
    lines.append("")
    if failing_factors or annex_status["FAIL"]:
        lines.append("The current spreadsheets should not yet be treated as API-verified source of truth. They contain table values and guardrail rows that require engineering review against the digitized API graph package.")
    else:
        lines.append("The current spreadsheets are broadly consistent with the digitized API graph package within the configured tolerance, subject to review of guardrail rows.")
    lines.append("")
    lines.append("## Files Created")
    lines.append("")
    created = [
        "figure_03_KHe.csv",
        "figure_04_Be.csv",
        "figure_05_Ee.csv",
        "figure_07_Fi.csv",
        "figure_08_KHr.csv",
        "figure_09_GHr.csv",
        "figure_10_NH.csv",
        "figure_11_KLr.csv",
        "figure_12_GLr.csv",
        "figure_13_NL.csv",
        "figure_14_KHh.csv",
        "figure_15_GHh.csv",
        "figure_16_KLh.csv",
        "figure_17_GLh.csv",
        "figure_18A_RF.csv",
        "figure_18B_RF.csv",
        "api1102_digitized_lookup_tables.xlsx",
        "spreadsheet_table_comparison.csv",
        "annex_b_control_point_check.csv",
        "calibration_metadata.json",
        "digitized_points_provenance.csv",
        "viewer_manifest.json",
        "graph_underlays/*.png",
        "overlays/*.png",
        "source_page_images/*.png",
    ]
    for name in created:
        lines.append(f"- `{name}`")
    lines.append("")
    (OUTPUT_DIR / "digitization_verification_report.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OVERLAY_DIR.mkdir(parents=True, exist_ok=True)
    PAGE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    CLEAN_GRAPH_DIR.mkdir(parents=True, exist_ok=True)

    spreadsheet_rows, spreadsheet_xs, workbook_tables = read_spreadsheet_points()
    doc = fitz.open(PDF_PATH)
    records_by_figure, dense_by_curve, metadata = generate_digitized_records(doc, spreadsheet_xs)
    write_raw_csvs(records_by_figure)
    write_lookup_workbook(records_by_figure)
    write_provenance(records_by_figure, metadata, doc)
    write_overlays(doc, records_by_figure)
    write_viewer_assets(doc, records_by_figure, metadata)
    comparison_rows = write_spreadsheet_comparison(spreadsheet_rows, dense_by_curve)
    annex_rows = write_annex_controls(dense_by_curve, workbook_tables)
    write_report(records_by_figure, metadata, comparison_rows, annex_rows)

    summary = {
        "figures": len(FIGURES),
        "curves": sum(len(spec.curves) for spec in FIGURES.values()),
        "points": sum(len(rows) for rows in records_by_figure.values()),
        "comparison_status": dict(Counter(row["status"] for row in comparison_rows)),
        "annex_status": dict(Counter(row["status"] for row in annex_rows)),
        "output_dir": str(OUTPUT_DIR),
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
