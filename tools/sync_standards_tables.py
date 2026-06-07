from __future__ import annotations

import math
import sys
from pathlib import Path
from pprint import pformat
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.api1102_digitization.digitize_api1102 import SPREADSHEET_SPECS


HIGHWAY_TABLES = ROOT / "app" / "standards" / "highway_tables.py"
RAILROAD_TABLES = ROOT / "app" / "standards" / "railroad_tables.py"
HIGHWAY_WORKBOOK = ROOT / "Refs" / "API 1102 Highway_260606.xlsx"
RAILROAD_WORKBOOK = ROOT / "Refs" / "API 1102 Railroad_260606.xlsx"
STANDARDS_VERSION = "workbook-sync-2026-06-07"


def numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result) or math.isinf(result):
        return None
    return result


def outer_key(factor: str, curve_name: str) -> float | str:
    if factor == "KHe":
        return {
            "E_prime_ksi=0.2": 200.0,
            "E_prime_ksi=0.5": 500.0,
            "E_prime_ksi=1.0": 1000.0,
            "E_prime_ksi=2.0": 2000.0,
        }[curve_name]
    if factor in {"KHh", "KHr", "KLh", "KLr"}:
        return {"Er_ksi=5": 5000.0, "Er_ksi=10": 10000.0, "Er_ksi=20": 20000.0}[curve_name]
    if factor in {"GHh", "GHr", "GLh", "GLr", "NH", "NL"}:
        return curve_name.replace("H_ft=", "").replace("_to_4", "")
    raise KeyError(f"{factor} {curve_name} has no outer key mapping.")


def empty_tables() -> dict[str, Any]:
    return {
        "EARTH_KHE_BY_E_PRIME": {},
        "BURIAL_A_BY_H_BD": {},
        "BURIAL_B_BY_H_BD": {},
        "EXCAVATION_BY_BD_D": {},
        "IMPACT_BY_COVER": {},
        "KH_BY_ER": {},
        "KL_BY_ER": {},
        "GH_BY_DEPTH": {},
        "GL_BY_DEPTH": {},
        "NH_BY_DEPTH": {},
        "NL_BY_DEPTH": {},
    }


def add_point(tables: dict[str, Any], factor: str, curve_name: str, x_value: float, y_value: float) -> None:
    if factor == "KHe":
        tables["EARTH_KHE_BY_E_PRIME"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor == "Be":
        name = "BURIAL_A_BY_H_BD" if curve_name == "soil_type=A" else "BURIAL_B_BY_H_BD"
        tables[name][x_value] = y_value
    elif factor == "Ee":
        tables["EXCAVATION_BY_BD_D"][x_value] = y_value
    elif factor == "Fi":
        tables["IMPACT_BY_COVER"][x_value] = y_value
    elif factor in {"KHh", "KHr"}:
        tables["KH_BY_ER"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor in {"KLh", "KLr"}:
        tables["KL_BY_ER"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor in {"GHh", "GHr"}:
        tables["GH_BY_DEPTH"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor in {"GLh", "GLr"}:
        tables["GL_BY_DEPTH"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor == "NH":
        tables["NH_BY_DEPTH"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    elif factor == "NL":
        tables["NL_BY_DEPTH"].setdefault(outer_key(factor, curve_name), {})[x_value] = y_value
    else:
        raise KeyError(f"Unsupported factor {factor}.")


def sort_table(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: sort_table(value[key]) for key in sorted(value)}
    return value


def read_workbook_tables() -> dict[str, dict[str, Any]]:
    tables_by_workbook = {"Highway": empty_tables(), "Railroad": empty_tables()}
    workbook_cache = {}
    for spec in SPREADSHEET_SPECS:
        workbook = workbook_cache.get(spec.path)
        if workbook is None:
            workbook = load_workbook(spec.path, data_only=True, read_only=True)
            workbook_cache[spec.path] = workbook
        sheet = workbook["Tables"]
        for row_index in range(spec.row_start, spec.row_end + 1):
            for pair in spec.pairs:
                x_value = numeric(sheet.cell(row_index, pair.x_col).value)
                y_value = numeric(sheet.cell(row_index, pair.y_col).value)
                if x_value is None or y_value is None:
                    continue
                add_point(tables_by_workbook[spec.workbook_label], spec.factor, pair.curve_name, x_value, y_value)
    for workbook in workbook_cache.values():
        workbook.close()
    return {mode: {name: sort_table(table) for name, table in tables.items()} for mode, tables in tables_by_workbook.items()}


def literal(value: Any) -> str:
    return pformat(value, width=120, sort_dicts=False)


def write_highway_tables(tables: dict[str, Any]) -> None:
    from app.standards.highway_tables import PAVEMENT_AXLE_FACTORS

    content = f'''"""Highway workbook lookup tables.

Generated from Refs/API 1102 Highway_260606.xlsx by tools/sync_standards_tables.py.
The nested tables use the existing two-step interpolation helper: outer key first,
inner lookup value second.
"""

EARTH_KHE_BY_E_PRIME = {literal(tables["EARTH_KHE_BY_E_PRIME"])}

BURIAL_A_BY_H_BD = {literal(tables["BURIAL_A_BY_H_BD"])}
BURIAL_B_BY_H_BD = {literal(tables["BURIAL_B_BY_H_BD"])}
EXCAVATION_BY_BD_D = {literal(tables["EXCAVATION_BY_BD_D"])}
IMPACT_BY_COVER = {literal(tables["IMPACT_BY_COVER"])}

KH_BY_ER = {literal(tables["KH_BY_ER"])}
KL_BY_ER = {literal(tables["KL_BY_ER"])}
GH_BY_DEPTH = {literal(tables["GH_BY_DEPTH"])}
GL_BY_DEPTH = {literal(tables["GL_BY_DEPTH"])}

PAVEMENT_AXLE_FACTORS = {literal(PAVEMENT_AXLE_FACTORS)}
'''
    HIGHWAY_TABLES.write_text(content, encoding="utf-8")


def write_railroad_tables(railroad: dict[str, Any], highway: dict[str, Any]) -> None:
    shared_names = ("EARTH_KHE_BY_E_PRIME", "BURIAL_A_BY_H_BD", "BURIAL_B_BY_H_BD", "EXCAVATION_BY_BD_D")
    shared_imports = [name for name in shared_names if railroad[name] == highway[name]]
    separate_names = [name for name in shared_names if railroad[name] != highway[name]]
    import_line = ""
    if shared_imports:
        import_line = f"from app.standards.highway_tables import {', '.join(shared_imports)}\n\n"
    shared_defs = "\n".join(f"{name} = {literal(railroad[name])}\n" for name in separate_names)
    content = f'''"""Railroad workbook lookup tables.

Generated from Refs/API 1102 Railroad_260606.xlsx by tools/sync_standards_tables.py.
"""

{import_line}{shared_defs}IMPACT_BY_COVER = {literal(railroad["IMPACT_BY_COVER"])}

KH_BY_ER = {literal(railroad["KH_BY_ER"])}
KL_BY_ER = {literal(railroad["KL_BY_ER"])}
GH_BY_DEPTH = {literal(railroad["GH_BY_DEPTH"])}
GL_BY_DEPTH = {literal(railroad["GL_BY_DEPTH"])}
NH_BY_DEPTH = {literal(railroad["NH_BY_DEPTH"])}
NL_BY_DEPTH = {literal(railroad["NL_BY_DEPTH"])}
'''
    RAILROAD_TABLES.write_text(content, encoding="utf-8")


def main() -> None:
    tables_by_workbook = read_workbook_tables()
    write_highway_tables(tables_by_workbook["Highway"])
    write_railroad_tables(tables_by_workbook["Railroad"], tables_by_workbook["Highway"])


if __name__ == "__main__":
    main()
